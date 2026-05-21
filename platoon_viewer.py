import html
import json
import re
import subprocess
import threading
import time
import unicodedata
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html.parser import HTMLParser
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional
from urllib.error import HTTPError, URLError
from urllib.parse import parse_qs, quote, unquote, urlparse
from urllib.request import Request, urlopen
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from fastapi import Body, FastAPI, HTTPException, Query
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.fangraphs.com/roster-resource/platoon-lineups"
BULLPEN_URL = "https://www.fangraphs.com/roster-resource/closer-depth-chart"
MLB_SCORES_URL = "https://www.mlb.com/scores"
SAVANT_PITCH_ARSENALS_URL = "https://baseballsavant.mlb.com/leaderboard/pitch-arsenals?year={year}&min={minimum}&type={metric_type}&hand="
APP_DIR = Path(__file__).resolve().parent
HOME_TEMPLATE_PATH = APP_DIR / "templates" / "index.html"
LINEUP_TEMPLATE_PATH = APP_DIR / "templates" / "platoon_lineups.html"
BULLPEN_TEMPLATE_PATH = APP_DIR / "templates" / "bullpens.html"
VELOCITY_TEMPLATE_PATH = APP_DIR / "templates" / "velocity_comparison.html"
PITCHERS_TEMPLATE_PATH = APP_DIR / "templates" / "probable_pitchers.html"
PITCHER_PROFILE_TEMPLATE_PATH = APP_DIR / "templates" / "pitcher_profile.html"
STARTING_PITCHERS_TEMPLATE_PATH = APP_DIR / "templates" / "starting_pitchers.html"
STUFF_PLUS_CACHE_PATH = APP_DIR / "pitch_stuff_plus_cache.json"
VELOCITY_COMPARISON_CACHE_PATH = APP_DIR / "velocity_comparison_cache.json"
LINEUP_CACHE_PATH = APP_DIR / "lineup_cache.json"
BULLPEN_CACHE_PATH = APP_DIR / "bullpen_cache.json"
PITCHER_STUFF_COMPARISON_CACHE_PATH = APP_DIR / "pitcher_stuff_comparison_cache.json"
HOMEPAGE_SCOREBOARD_CACHE_PATH = APP_DIR / "homepage_scoreboard_cache.json"
STARTING_PITCHER_OVERRIDES_PATH = APP_DIR / "starting_pitcher_overrides.json"
STARTING_PITCHER_PREVIEW_CACHE_PATH = APP_DIR / "starting_pitcher_preview_cache.json"
PITCHER_PROFILE_PAYLOAD_CACHE_PATH = APP_DIR / "pitcher_profile_payload_cache.json"
MLB_PRICING_WORKBOOK_PATH = APP_DIR / "MLB Pricing.xlsx"
CHROME_PROFILE_DIR = APP_DIR / "chrome_fangraphs_profile"
CHROME_PATH_CANDIDATES = [
    Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
    Path.home() / r"AppData\Local\Google\Chrome\Application\chrome.exe",
]
REFRESH_SECONDS = 300
VELOCITY_RUNTIME_CACHE_SECONDS = 300
HOMEPAGE_SCOREBOARD_CACHE_SECONDS = 300
HOMEPAGE_LINEUP_CACHE_SECONDS = 900
PITCHER_PROFILE_CACHE_SECONDS = 900
MLB_HEADSHOT_TEMPLATE = "https://midfield.mlbstatic.com/v1/people/{mlb_id}/silo/360?zoom=1.2"
VALID_BATS = {"R", "L", "S"}
POSITIONS = {
    "C",
    "1B",
    "2B",
    "3B",
    "SS",
    "LF",
    "CF",
    "RF",
    "DH",
    "INF",
    "OF",
    "UT",
    "UTIL",
    "1B/OF",
    "2B/SS",
    "3B/SS",
    "INF/OF",
    "OF/INF",
}
STARTING_LINEUP_LENGTH = 9
TEAM_OPTIONS = [
    {"slug": "diamondbacks", "code": "ARI", "name": "Arizona Diamondbacks", "league": "NL West"},
    {"slug": "braves", "code": "ATL", "name": "Atlanta Braves", "league": "NL East"},
    {"slug": "orioles", "code": "BAL", "name": "Baltimore Orioles", "league": "AL East"},
    {"slug": "red-sox", "code": "BOS", "name": "Boston Red Sox", "league": "AL East"},
    {"slug": "cubs", "code": "CHC", "name": "Chicago Cubs", "league": "NL Central"},
    {"slug": "white-sox", "code": "CHW", "name": "Chicago White Sox", "league": "AL Central"},
    {"slug": "reds", "code": "CIN", "name": "Cincinnati Reds", "league": "NL Central"},
    {"slug": "guardians", "code": "CLE", "name": "Cleveland Guardians", "league": "AL Central"},
    {"slug": "rockies", "code": "COL", "name": "Colorado Rockies", "league": "NL West"},
    {"slug": "tigers", "code": "DET", "name": "Detroit Tigers", "league": "AL Central"},
    {"slug": "astros", "code": "HOU", "name": "Houston Astros", "league": "AL West"},
    {"slug": "royals", "code": "KCR", "name": "Kansas City Royals", "league": "AL Central"},
    {"slug": "angels", "code": "LAA", "name": "Los Angeles Angels", "league": "AL West"},
    {"slug": "dodgers", "code": "LAD", "name": "Los Angeles Dodgers", "league": "NL West"},
    {"slug": "marlins", "code": "MIA", "name": "Miami Marlins", "league": "NL East"},
    {"slug": "brewers", "code": "MIL", "name": "Milwaukee Brewers", "league": "NL Central"},
    {"slug": "twins", "code": "MIN", "name": "Minnesota Twins", "league": "AL Central"},
    {"slug": "mets", "code": "NYM", "name": "New York Mets", "league": "NL East"},
    {"slug": "yankees", "code": "NYY", "name": "New York Yankees", "league": "AL East"},
    {"slug": "athletics", "code": "OAK", "name": "Athletics", "league": "AL West"},
    {"slug": "phillies", "code": "PHI", "name": "Philadelphia Phillies", "league": "NL East"},
    {"slug": "pirates", "code": "PIT", "name": "Pittsburgh Pirates", "league": "NL Central"},
    {"slug": "padres", "code": "SDP", "name": "San Diego Padres", "league": "NL West"},
    {"slug": "giants", "code": "SFG", "name": "San Francisco Giants", "league": "NL West"},
    {"slug": "mariners", "code": "SEA", "name": "Seattle Mariners", "league": "AL West"},
    {"slug": "cardinals", "code": "STL", "name": "St. Louis Cardinals", "league": "NL Central"},
    {"slug": "rays", "code": "TBR", "name": "Tampa Bay Rays", "league": "AL East"},
    {"slug": "rangers", "code": "TEX", "name": "Texas Rangers", "league": "AL West"},
    {"slug": "blue-jays", "code": "TOR", "name": "Toronto Blue Jays", "league": "AL East"},
    {"slug": "nationals", "code": "WSN", "name": "Washington Nationals", "league": "NL East"},
]
TEAM_MAP = {team["slug"]: team for team in TEAM_OPTIONS}
MLB_TEAM_NAME_TO_SLUG = {
    "arizona": "diamondbacks",
    "d-backs": "diamondbacks",
    "atlanta": "braves",
    "baltimore": "orioles",
    "boston": "red-sox",
    "chi cubs": "cubs",
    "chicago cubs": "cubs",
    "chi white sox": "white-sox",
    "chicago white sox": "white-sox",
    "cincinnati": "reds",
    "cleveland": "guardians",
    "colorado": "rockies",
    "detroit": "tigers",
    "houston": "astros",
    "kansas city": "royals",
    "la angels": "angels",
    "los angeles angels": "angels",
    "la dodgers": "dodgers",
    "los angeles dodgers": "dodgers",
    "miami": "marlins",
    "milwaukee": "brewers",
    "minnesota": "twins",
    "ny mets": "mets",
    "new york mets": "mets",
    "ny yankees": "yankees",
    "new york yankees": "yankees",
    "athletics": "athletics",
    "philadelphia": "phillies",
    "pittsburgh": "pirates",
    "san diego": "padres",
    "san francisco": "giants",
    "seattle": "mariners",
    "st. louis": "cardinals",
    "st louis": "cardinals",
    "tampa bay": "rays",
    "texas": "rangers",
    "toronto": "blue-jays",
    "washington": "nationals",
}
DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Cache-Control": "max-age=0",
    "Upgrade-Insecure-Requests": "1",
}
FANGRAPHS_HEADERS = {
    **DEFAULT_HEADERS,
    "Referer": "https://www.fangraphs.com/",
}
JSON_HEADERS = {
    "User-Agent": DEFAULT_HEADERS["User-Agent"],
    "Accept": "application/json",
}
FANGRAPHS_URL_CACHE: Dict[str, Optional[str]] = {}
PITCHER_PROFILE_CACHE: Dict[int, Dict[str, object]] = {}
DEPTH_CHART_CACHE: Dict[str, List[Dict[str, object]]] = {}
DEPTH_CHART_DATA_CACHE: Dict[str, Dict[str, object]] = {}
FANGRAPHS_PITCHER_ID_MAP_CACHE: Optional[Dict[str, int]] = None
PITCH_STUFF_CACHE: Dict[str, Dict[str, object]] = {}
HOMEPAGE_PITCHER_HAND_CACHE: Dict[int, str] = {}
VELOCITY_RUNTIME_CACHE: Dict[str, object] = {}
VELOCITY_RUNTIME_CACHE_LOCK = threading.Lock()
STARTING_PITCHER_RUNTIME_CACHE: Dict[str, object] = {}
STARTING_PITCHER_RUNTIME_CACHE_LOCK = threading.Lock()
HOMEPAGE_SCOREBOARD_RUNTIME_CACHE: Dict[str, object] = {}
HOMEPAGE_SCOREBOARD_RUNTIME_CACHE_LOCK = threading.Lock()
STARTING_PITCHER_PREVIEW_RUNTIME_CACHE: Dict[str, object] = {}
STARTING_PITCHER_PREVIEW_RUNTIME_CACHE_LOCK = threading.Lock()
HOMEPAGE_CACHE_REFRESH_THREAD_STARTED = False
HOMEPAGE_CACHE_REFRESH_THREAD_LOCK = threading.Lock()
STUFF_PLUS_REFRESH_LOCK = threading.Lock()
STUFF_PLUS_REFRESH_IN_PROGRESS = False
SAVANT_PITCH_TYPES = [
    ("ff", "Four-Seam"),
    ("si", "Sinker"),
    ("fc", "Cutter"),
    ("fs", "Splitter"),
    ("fo", "Forkball"),
    ("sl", "Slider"),
    ("st", "Sweeper"),
    ("cu", "Curveball"),
    ("sv", "Slurve"),
    ("ch", "Changeup"),
    ("sc", "Screwball"),
    ("kn", "Knuckleball"),
]
STUFF_PLUS_FIELD_MAP = {
    "FF": "FF",
    "SI": "SI",
    "FC": "FC",
    "FS": "FS",
    "FO": "FO",
    "SL": "SL",
    "ST": "SL",
    "SV": "SL",
    "CU": "CU",
    "CH": "CH",
    "KC": "KC",
    "KN": None,
    "SC": None,
}
@dataclass
class PlayerRow:
    status: str
    position: str
    name: str
    bats: str
    stats: List[str]


@dataclass
class BullpenRow:
    name: str
    throws: str
    role: str
    usage: List[str]
    stats: List[str]


@dataclass
class BullpenInjuryRow:
    name: str
    throws: str
    status: str


class LineBreakHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: List[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() in {"h1", "h2", "h3", "div", "p", "br", "tr", "li", "section", "header"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"h1", "h2", "h3", "div", "p", "tr", "li", "section", "header"}:
            self.parts.append("\n")
        elif tag.lower() in {"td", "th", "a", "span"}:
            self.parts.append(" ")

    def handle_data(self, data: str) -> None:
        if data:
            self.parts.append(data)
            self.parts.append(" ")

    def get_text(self) -> str:
        return "".join(self.parts)


def resolve_chrome_path() -> Optional[Path]:
    for candidate in CHROME_PATH_CANDIDATES:
        if candidate.exists():
            return candidate
    return None


def fetch_url(url: str) -> bytes:
    req = Request(url, headers=DEFAULT_HEADERS)
    with urlopen(req, timeout=30) as resp:
        return resp.read()


def fetch_json(url: str) -> Dict[str, object]:
    req = Request(url, headers=JSON_HEADERS)
    with urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8", "ignore"))


def fetch_mlb_person(mlb_player_id: int) -> Dict[str, object]:
    payload = fetch_json(
        f"https://statsapi.mlb.com/api/v1/people/{mlb_player_id}?hydrate=currentTeam"
    )
    people = payload.get("people", [])
    if not people:
        raise ValueError("Could not find MLB player bio.")
    person = people[0]
    if not isinstance(person, dict):
        raise ValueError("Could not parse MLB player bio.")
    return person


def fetch_fangraphs_page(url: str) -> str:
    req = Request(url, headers=FANGRAPHS_HEADERS)
    try:
        with urlopen(req, timeout=30) as resp:
            return resp.read().decode("utf-8", "ignore")
    except Exception as exc:
        chrome_path = resolve_chrome_path()
        if not chrome_path:
            raise exc
        return fetch_fangraphs_page_via_chrome(url, chrome_path)


def fetch_fangraphs_page_via_chrome(url: str, chrome_path: Path) -> str:
    CHROME_PROFILE_DIR.mkdir(exist_ok=True)
    command = [
        str(chrome_path),
        f"--user-data-dir={CHROME_PROFILE_DIR}",
        "--headless=new",
        "--disable-gpu",
        "--disable-blink-features=AutomationControlled",
        "--window-size=1400,2200",
        "--virtual-time-budget=8000",
        "--dump-dom",
        url,
    ]
    completed = subprocess.run(
        command,
        capture_output=True,
        text=True,
        timeout=45,
        check=False,
    )
    html_text = completed.stdout or ""
    if completed.returncode != 0 or not html_text.strip():
        stderr = (completed.stderr or "").strip()
        raise ValueError(f"Chrome fallback failed to fetch Fangraphs page: {stderr or 'empty response'}")
    return html_text


def fetch_text_page(url: str) -> str:
    req = Request(url, headers=DEFAULT_HEADERS)
    with urlopen(req, timeout=30) as resp:
        return resp.read().decode("utf-8", "ignore")


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").lower())


def resolve_team_slug(team_name: str) -> Optional[str]:
    return MLB_TEAM_NAME_TO_SLUG.get((team_name or "").strip().lower())


def next_mlb_date() -> date:
    return eastern_now().date() + timedelta(days=1)


def extract_lines(raw_html: str) -> List[str]:
    parser = LineBreakHTMLParser()
    parser.feed(raw_html)
    text = html.unescape(parser.get_text()).replace("\xa0", " ")
    lines = []
    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line).strip()
        if line:
            lines.append(line)
    return lines


def looks_like_game_time(line: str) -> bool:
    return bool(re.match(r"^\d{1,2}:\d{2}\s+[AP]M ET$", line))


def looks_like_pitcher_line(line: str) -> bool:
    return " HP " in line or line == "TBD" or " | " in line


def looks_like_schedule_note(line: str) -> bool:
    normalized = (line or "").strip().lower()
    if not normalized:
        return False
    return normalized.startswith("makeup of ") or normalized.startswith("completion of ")


def clean_probable_pitcher(line: str) -> Dict[str, str]:
    cleaned = re.sub(r"\s+", " ", line).strip()
    if cleaned == "TBD":
        return {"name": "TBD", "hand": "", "record": "", "era": ""}

    match = re.match(
        r"^(?P<name>.+?)\s+(?P<hand>[RLS])\s+HP(?:\s+(?P<record>\d+\s*-\s*\d+))?(?:\s*\|\s*(?P<era>.+))?$",
        cleaned,
    )
    if match:
        return {
            "name": match.group("name").strip(),
            "hand": match.group("hand") or "",
            "record": (match.group("record") or "").strip(),
            "era": (match.group("era") or "").strip(),
        }

    return {"name": cleaned, "hand": "", "record": "", "era": ""}


def extract_pitcher_entries(lines: List[str]) -> List[Dict[str, str]]:
    entries: List[Dict[str, str]] = []
    pattern = re.compile(
        r"(?P<name>.+?)\s+(?P<hand>[RLS])\s+HP(?:\s+(?P<record>\d+\s*-\s*\d+))?(?:\s*\|\s*(?P<era>\d+(?:\.\d+)?\s+ERA))?"
    )
    hand_line_pattern = re.compile(r"^(?P<hand>[RLS])\s+HP$")
    stat_line_pattern = re.compile(r"^(?P<record>\d+\s*-\s*\d+)(?:\s*\|\s*(?P<era>\d+(?:\.\d+)?\s+ERA))?$")

    normalized_lines = [re.sub(r"\s+", " ", raw_line).strip() for raw_line in lines]
    idx = 0
    while idx < len(normalized_lines):
        line = normalized_lines[idx]
        if not line or line == "-":
            idx += 1
            continue
        if line == "TBD":
            entries.append({"name": "TBD", "hand": "", "record": "", "era": ""})
            idx += 1
            if idx < len(normalized_lines) and normalized_lines[idx] == "-":
                idx += 1
            continue

        if (
            idx + 2 < len(normalized_lines)
            and hand_line_pattern.match(normalized_lines[idx + 1])
            and stat_line_pattern.match(normalized_lines[idx + 2])
        ):
            hand_match = hand_line_pattern.match(normalized_lines[idx + 1])
            stat_match = stat_line_pattern.match(normalized_lines[idx + 2])
            entries.append(
                {
                    "name": line,
                    "hand": hand_match.group("hand"),
                    "record": stat_match.group("record").strip(),
                    "era": (stat_match.group("era") or "").strip(),
                }
            )
            idx += 3
            continue

        matches = list(pattern.finditer(line))
        if matches:
            for match in matches:
                entries.append(
                    {
                        "name": match.group("name").strip(),
                        "hand": match.group("hand") or "",
                        "record": (match.group("record") or "").strip(),
                        "era": (match.group("era") or "").strip(),
                    }
                )
            idx += 1
            continue

        if looks_like_pitcher_line(line):
            entries.append(clean_probable_pitcher(line))
        idx += 1

    return entries


def extract_mlb_pitcher_cards(raw_html: str) -> List[Dict[str, str]]:
    pattern = re.compile(
        r'<a href="/player/[^"]+-(?P<mlb_id>\d+)"[^>]*>.*?<div name="(?P<display_name>[^"]+)" class="playerMatchupstyle__PlayerMatchupWrapper.*?<img[^>]+src="(?P<headshot_url>[^"]+)"',
        re.DOTALL,
    )
    cards: List[Dict[str, str]] = []
    for match in pattern.finditer(raw_html):
        cards.append(
            {
                "mlb_id": match.group("mlb_id"),
                "display_name": html.unescape(match.group("display_name")),
                "headshot_url": html.unescape(match.group("headshot_url")),
            }
        )
    return cards


def attach_pitcher_card_details(games: List[Dict[str, object]], cards: List[Dict[str, str]]) -> None:
    card_index = 0
    for game in games:
        for side in ("away_pitcher", "home_pitcher"):
            pitcher = game[side]
            if pitcher.get("name") == "TBD":
                pitcher["mlb_id"] = None
                pitcher["headshot_url"] = ""
                continue
            if card_index >= len(cards):
                pitcher["mlb_id"] = None
                pitcher["headshot_url"] = ""
                continue

            card = cards[card_index]
            card_index += 1
            pitcher["mlb_id"] = int(card["mlb_id"])
            pitcher["headshot_url"] = card["headshot_url"]


def parse_probable_pitchers_page(target_date: date, raw_html: str) -> Dict[str, object]:
    lines = extract_lines(raw_html)
    games: List[Dict[str, object]] = []
    pitcher_cards = extract_mlb_pitcher_cards(raw_html)
    idx = 0

    while idx < len(lines):
        line = lines[idx]
        if not looks_like_game_time(line):
            idx += 1
            continue

        game_time = line
        search_end = min(idx + 40, len(lines))
        preview_index = None
        for look_ahead in range(idx + 1, search_end):
            if lines[look_ahead] == "Preview":
                preview_index = look_ahead
                break

        if preview_index is None:
            idx += 1
            continue

        window = lines[idx:preview_index]
        watch_index = next((i for i, value in enumerate(window) if value.startswith("Watch on:")), None)
        if watch_index is None or watch_index + 4 >= len(window):
            idx = preview_index + 1
            continue

        post_watch_lines = [value for value in window[watch_index + 1:] if value and not looks_like_schedule_note(value)]
        if len(post_watch_lines) < 2:
            idx = preview_index + 1
            continue

        away_team = post_watch_lines[0]
        home_team = post_watch_lines[1]
        pitcher_entries = extract_pitcher_entries(post_watch_lines[2:])
        if not pitcher_entries:
            idx = preview_index + 1
            continue

        if len(pitcher_entries) == 1:
            away_pitcher = pitcher_entries[0]
            home_pitcher = {"name": "TBD", "hand": "", "record": "", "era": ""}
        else:
            away_pitcher = pitcher_entries[0]
            home_pitcher = pitcher_entries[1]

        games.append(
            {
                "game_time": game_time,
                "away_team": away_team,
                "home_team": home_team,
                "away_pitcher": away_pitcher,
                "home_pitcher": home_pitcher,
            }
        )
        idx = preview_index + 1

    attach_pitcher_card_details(games, pitcher_cards)
    attach_fangraphs_ids(target_date, games)

    return {
        "date": target_date.isoformat(),
        "source_url": f"{MLB_SCORES_URL}/{target_date.isoformat()}",
        "games": games,
        "default_date": next_mlb_date().isoformat(),
    }


def fetch_probable_pitchers(target_date: date) -> Dict[str, object]:
    raw = fetch_url(f"{MLB_SCORES_URL}/{target_date.isoformat()}")
    html_text = raw.decode("utf-8", "ignore")
    return parse_probable_pitchers_page(target_date, html_text)


def format_schedule_game_time(game_datetime: Optional[str]) -> str:
    if not game_datetime:
        return "Time TBD"
    try:
        normalized = game_datetime.replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
        if parsed.tzinfo is None:
            return parsed.strftime("%I:%M %p ET").lstrip("0")
        try:
            eastern = parsed.astimezone(ZoneInfo("America/New_York"))
        except ZoneInfoNotFoundError:
            eastern = parsed.astimezone()
        return eastern.strftime("%I:%M %p ET").lstrip("0")
    except Exception:
        return str(game_datetime)


def normalize_pitcher_lookup_name(name: object) -> str:
    if not name:
        return ""
    normalized = unicodedata.normalize("NFKD", str(name))
    ascii_name = normalized.encode("ascii", "ignore").decode("ascii")
    collapsed = re.sub(r"[^a-z0-9]+", "", ascii_name.lower())
    return collapsed


def build_starting_pitcher_preview_lookup() -> Dict[str, Dict[str, object]]:
    try:
        return get_or_build_starting_pitcher_preview_lookup()
    except Exception:
        return {}


def display_starting_pitcher_header(header: str) -> str:
    return "2025 Stuff+" if header == "2025 Stuff+2" else header


def build_starting_pitcher_profile_section(pitcher_name: str) -> Dict[str, object]:
    try:
        payload = build_starting_pitcher_payload()
    except Exception:
        return {"matched": False, "fields": [], "sheet_name": "SPs"}

    normalized_target = normalize_pitcher_lookup_name(pitcher_name)
    candidate_row: Optional[Dict[str, object]] = None
    surname_matches: List[Dict[str, object]] = []

    for row in payload.get("rows", []):
        if not isinstance(row, dict):
            continue
        row_name = str(row.get("Pitcher") or "").strip()
        if not row_name:
            continue
        normalized_row_name = normalize_pitcher_lookup_name(row_name)
        if normalized_row_name == normalized_target:
            candidate_row = row
            break
        tokens = re.findall(r"[A-Za-z0-9']+", row_name)
        if tokens and normalize_pitcher_lookup_name(tokens[-1]) == normalized_target:
            surname_matches.append(row)

    if candidate_row is None and len(surname_matches) == 1:
        candidate_row = surname_matches[0]

    headers = [header for header in payload.get("headers", []) if header != "Pitcher"]
    if not candidate_row:
        return {
            "matched": False,
            "sheet_name": payload.get("source", {}).get("sheet_name", "SPs"),
            "fields": [{"key": header, "label": display_starting_pitcher_header(header), "value": ""} for header in headers],
        }

    return {
        "matched": True,
        "sheet_name": payload.get("source", {}).get("sheet_name", "SPs"),
        "fields": [
            {
                "key": header,
                "label": display_starting_pitcher_header(header),
                "value": candidate_row.get(header, ""),
            }
            for header in headers
        ],
    }


def enrich_scoreboard_pitcher_from_sp_board(
    pitcher_payload: Dict[str, object],
    sp_lookup: Dict[str, Dict[str, object]],
) -> Dict[str, object]:
    if not pitcher_payload:
        return pitcher_payload

    lookup_key = normalize_pitcher_lookup_name(pitcher_payload.get("name"))
    preview = sp_lookup.get(lookup_key, {}) if lookup_key else {}
    if not preview:
        return pitcher_payload

    enriched = dict(pitcher_payload)
    era_value = preview.get("era")
    ip_value = preview.get("ip")
    if era_value not in (None, ""):
        enriched["era"] = f"{era_value} ERA"
    if ip_value not in (None, ""):
        enriched["ip"] = f"{ip_value} IP"
    return enriched


def build_scoreboard_pitcher_payload(
    probable_pitcher: Optional[Dict[str, object]],
    sp_lookup: Optional[Dict[str, Dict[str, object]]] = None,
) -> Dict[str, object]:
    probable_pitcher = probable_pitcher or {}
    payload = {
        "name": probable_pitcher.get("fullName") or "TBD",
        "mlb_id": probable_pitcher.get("id"),
        "hand": "",
        "record": "",
        "era": "",
        "ip": "",
    }
    if sp_lookup:
        return enrich_scoreboard_pitcher_from_sp_board(payload, sp_lookup)
    return payload


def parse_cached_iso_datetime(value: object) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        return None


def cache_timestamp_is_fresh(timestamp_value: object, ttl_seconds: int) -> bool:
    parsed = parse_cached_iso_datetime(timestamp_value)
    if not parsed:
        return False
    return (eastern_now() - parsed).total_seconds() <= ttl_seconds


def get_cached_homepage_lineup_fallback(
    homepage_cache: Dict[str, object],
    team_slug: Optional[str],
    opposing_hand: str,
) -> Optional[Dict[str, object]]:
    if not team_slug:
        return None
    lineup_fallbacks = homepage_cache.get("lineup_fallbacks", {}) if isinstance(homepage_cache, dict) else {}
    if not isinstance(lineup_fallbacks, dict):
        return None
    hand_key = "vsL" if opposing_hand == "L" else "vsR"
    cache_key = f"{team_slug}:{hand_key}"
    cached_entry = lineup_fallbacks.get(cache_key)
    if not isinstance(cached_entry, dict):
        return None
    if not cache_timestamp_is_fresh(cached_entry.get("saved_at"), HOMEPAGE_LINEUP_CACHE_SECONDS):
        return None
    lineup_payload = cached_entry.get("lineup")
    return dict(lineup_payload) if isinstance(lineup_payload, dict) else None


def save_cached_homepage_lineup_fallback(
    homepage_cache: Dict[str, object],
    team_slug: Optional[str],
    opposing_hand: str,
    lineup_payload: Dict[str, object],
) -> None:
    if not team_slug:
        return
    lineup_fallbacks = homepage_cache.setdefault("lineup_fallbacks", {}) if isinstance(homepage_cache, dict) else {}
    if not isinstance(lineup_fallbacks, dict):
        lineup_fallbacks = {}
        homepage_cache["lineup_fallbacks"] = lineup_fallbacks
    hand_key = "vsL" if opposing_hand == "L" else "vsR"
    cache_key = f"{team_slug}:{hand_key}"
    lineup_fallbacks[cache_key] = {
        "saved_at": eastern_now().isoformat(),
        "lineup": lineup_payload,
    }


def get_pitcher_hand(mlb_player_id: Optional[int]) -> str:
    if not mlb_player_id:
        return ""
    if mlb_player_id in HOMEPAGE_PITCHER_HAND_CACHE:
        return HOMEPAGE_PITCHER_HAND_CACHE[mlb_player_id]
    try:
        person = fetch_mlb_person(mlb_player_id)
        code = str((person.get("pitchHand") or {}).get("code") or "").upper()
        if code not in {"R", "L"}:
            code = ""
    except Exception:
        code = ""
    HOMEPAGE_PITCHER_HAND_CACHE[mlb_player_id] = code
    return code


def build_homepage_lineup_from_mlb_players(players: List[Dict[str, object]]) -> Dict[str, object]:
    return {
        "source": "official",
        "is_official": True,
        "players": [
            {
                "name": player.get("fullName") or "",
                "position": ((player.get("primaryPosition") or {}).get("abbreviation") or ""),
                "confirmed": True,
            }
            for player in players[:STARTING_LINEUP_LENGTH]
            if player.get("fullName")
        ],
    }


def build_homepage_lineup_from_fangraphs(
    team_slug: Optional[str],
    opposing_hand: str,
    homepage_cache: Optional[Dict[str, object]] = None,
) -> Dict[str, object]:
    if not team_slug:
        return {"source": "fangraphs", "is_official": False, "players": []}
    cached_lineup = get_cached_homepage_lineup_fallback(homepage_cache or {}, team_slug, opposing_hand)
    if cached_lineup:
        return cached_lineup
    try:
        payload = fetch_lineup_data(team_slug)
        split_key = "vsL" if opposing_hand == "L" else "vsR"
        rows = (payload.get("sections", {}) or {}).get(split_key, {}).get("rows", [])[:STARTING_LINEUP_LENGTH]
        lineup_payload = {
            "source": f"fangraphs_{split_key}",
            "is_official": False,
            "players": [
                {
                    "name": row.get("name") or "",
                    "position": row.get("position") or "",
                    "confirmed": False,
                }
                for row in rows
                if row.get("name")
            ],
        }
        if homepage_cache is not None:
            save_cached_homepage_lineup_fallback(homepage_cache, team_slug, opposing_hand, lineup_payload)
        return lineup_payload
    except Exception:
        return cached_lineup or {"source": "fangraphs", "is_official": False, "players": []}


def build_homepage_team_lineup(
    team_name: str,
    official_players: Optional[List[Dict[str, object]]],
    opposing_pitcher_id: Optional[int],
    homepage_cache: Optional[Dict[str, object]] = None,
) -> Dict[str, object]:
    if official_players:
        return build_homepage_lineup_from_mlb_players(official_players)
    team_slug = resolve_team_slug(team_name)
    opposing_hand = get_pitcher_hand(opposing_pitcher_id)
    return build_homepage_lineup_from_fangraphs(team_slug, opposing_hand, homepage_cache)


def is_homepage_game_pregame(status: Dict[str, object]) -> bool:
    abstract = str((status or {}).get("abstract") or "").strip().lower()
    detailed = str((status or {}).get("detailed") or "").strip().lower()
    return abstract == "preview" or detailed in {"scheduled", "pre-game", "warmup", "delayed start"}


def build_homepage_game_payload(
    game: Dict[str, object],
    sp_lookup: Optional[Dict[str, Dict[str, object]]] = None,
    homepage_cache: Optional[Dict[str, object]] = None,
) -> Dict[str, object]:
    away = game.get("teams", {}).get("away", {})
    home = game.get("teams", {}).get("home", {})
    away_team = away.get("team", {}) or {}
    home_team = home.get("team", {}) or {}
    status = game.get("status", {}) or {}
    broadcasts = game.get("broadcasts", []) or []
    lineups = game.get("lineups", {}) or {}
    broadcast_names = [
        entry.get("name")
        for entry in broadcasts
        if isinstance(entry, dict) and entry.get("name")
    ]
    away_pitcher = build_scoreboard_pitcher_payload(away.get("probablePitcher"), sp_lookup)
    home_pitcher = build_scoreboard_pitcher_payload(home.get("probablePitcher"), sp_lookup)

    return {
        "game_pk": game.get("gamePk"),
        "status": {
            "abstract": status.get("abstractGameState") or "",
            "detailed": status.get("detailedState") or "",
        },
        "game_time": format_schedule_game_time(game.get("gameDate")),
        "away_team": away_team.get("shortName") or away_team.get("name") or "",
        "home_team": home_team.get("shortName") or home_team.get("name") or "",
        "away_record": "{wins}-{losses}".format(
            wins=(away.get("leagueRecord") or {}).get("wins", 0),
            losses=(away.get("leagueRecord") or {}).get("losses", 0),
        ),
        "home_record": "{wins}-{losses}".format(
            wins=(home.get("leagueRecord") or {}).get("wins", 0),
            losses=(home.get("leagueRecord") or {}).get("losses", 0),
        ),
        "broadcasts": broadcast_names,
        "away_pitcher": away_pitcher,
        "home_pitcher": home_pitcher,
        "away_lineup": build_homepage_team_lineup(
            away_team.get("shortName") or away_team.get("name") or "",
            lineups.get("awayPlayers") if isinstance(lineups.get("awayPlayers"), list) else None,
            home_pitcher.get("mlb_id"),
            homepage_cache,
        ),
        "home_lineup": build_homepage_team_lineup(
            home_team.get("shortName") or home_team.get("name") or "",
            lineups.get("homePlayers") if isinstance(lineups.get("homePlayers"), list) else None,
            away_pitcher.get("mlb_id"),
            homepage_cache,
        ),
    }


def normalize_homepage_game_to_prematch(game_payload: Dict[str, object]) -> Dict[str, object]:
    normalized = dict(game_payload)
    normalized["status"] = {
        "abstract": "Preview",
        "detailed": "Scheduled",
    }
    return normalized


def fetch_homepage_slate(target_date: date, label: str) -> Dict[str, object]:
    cache_date_key = target_date.isoformat()
    homepage_cache = load_homepage_scoreboard_cache_file()
    cached_slates = homepage_cache.get("slates", {}) if isinstance(homepage_cache, dict) else {}
    cached_slate = cached_slates.get(cache_date_key, {}) if isinstance(cached_slates, dict) else {}
    if isinstance(cached_slate, dict) and cache_timestamp_is_fresh(
        cached_slate.get("generated_at"), HOMEPAGE_SCOREBOARD_CACHE_SECONDS
    ):
        return dict(cached_slate)

    schedule_url = (
        "https://statsapi.mlb.com/api/v1/schedule"
        f"?sportId=1&date={target_date.isoformat()}&hydrate=probablePitcher,lineups,decisions,linescore,team,flags,broadcasts(all)"
    )
    payload = fetch_json(schedule_url)
    date_entries = payload.get("dates", [])
    game_entries = date_entries[0].get("games", []) if date_entries else []
    cached_games_by_pk = {
        str(game.get("game_pk")): game
        for game in cached_slate.get("games", [])
        if isinstance(game, dict) and game.get("game_pk") is not None
    }
    games: List[Dict[str, object]] = []
    sp_lookup = build_starting_pitcher_preview_lookup()

    for game in game_entries:
        built_game = build_homepage_game_payload(game, sp_lookup, homepage_cache)
        game_pk = str(built_game.get("game_pk"))
        live_status = built_game.get("status", {})
        frozen_game = normalize_homepage_game_to_prematch(built_game)
        cached_game = cached_games_by_pk.get(game_pk)
        if cached_game:
            cached_game = {
                **cached_game,
                **frozen_game,
            }
        if is_homepage_game_pregame(live_status):
            cached_games_by_pk[game_pk] = frozen_game
            games.append(frozen_game)
        elif cached_game:
            cached_games_by_pk[game_pk] = cached_game
            games.append(cached_game)
        else:
            cached_games_by_pk[game_pk] = frozen_game
            games.append(frozen_game)

    slate_payload = {
        "label": label,
        "date": cache_date_key,
        "generated_at": eastern_now().isoformat(),
        "source_url": f"{MLB_SCORES_URL}/{cache_date_key}",
        "games": games,
        "error": None,
    }
    cached_slates[cache_date_key] = slate_payload
    homepage_cache["saved_at"] = eastern_now().isoformat()
    homepage_cache["slates"] = cached_slates
    save_homepage_scoreboard_cache_file(homepage_cache)
    return slate_payload


def build_homepage_scoreboard_payload(day_key: Optional[str] = None) -> Dict[str, object]:
    today_date = eastern_now().date()
    tomorrow_date = today_date + timedelta(days=1)
    runtime_cache_key = json.dumps(
        {
            "day": day_key or "all",
            "today": today_date.isoformat(),
            "tomorrow": tomorrow_date.isoformat(),
        },
        sort_keys=True,
    )
    current_timestamp = time.time()

    with HOMEPAGE_SCOREBOARD_RUNTIME_CACHE_LOCK:
        if (
            HOMEPAGE_SCOREBOARD_RUNTIME_CACHE.get("cache_key") == runtime_cache_key
            and HOMEPAGE_SCOREBOARD_RUNTIME_CACHE.get("expires_at", 0) > current_timestamp
            and HOMEPAGE_SCOREBOARD_RUNTIME_CACHE.get("payload")
        ):
            return dict(HOMEPAGE_SCOREBOARD_RUNTIME_CACHE["payload"])

    def load_slate(target_date: date, label: str) -> Dict[str, object]:
        try:
            return fetch_homepage_slate(target_date, label)
        except Exception as exc:
            return {
                "label": label,
                "date": target_date.isoformat(),
                "generated_at": eastern_now().isoformat(),
                "source_url": f"{MLB_SCORES_URL}/{target_date.isoformat()}",
                "games": [],
                "error": str(exc),
            }

    payload: Dict[str, object] = {
        "generated_at": eastern_now().isoformat(),
        "available_days": ["today", "tomorrow"],
    }
    if day_key in (None, "today"):
        payload["today"] = load_slate(today_date, "Today")
    if day_key in (None, "tomorrow"):
        payload["tomorrow"] = load_slate(tomorrow_date, "Tomorrow")

    with HOMEPAGE_SCOREBOARD_RUNTIME_CACHE_LOCK:
        HOMEPAGE_SCOREBOARD_RUNTIME_CACHE.clear()
        HOMEPAGE_SCOREBOARD_RUNTIME_CACHE.update(
            {
                "cache_key": runtime_cache_key,
                "expires_at": current_timestamp + 60,
                "payload": payload,
            }
        )
    return payload


def refresh_homepage_supporting_caches_once() -> None:
    try:
        get_or_build_starting_pitcher_preview_lookup(force_refresh=True)
    except Exception:
        pass

    today_date = eastern_now().date()
    tomorrow_date = today_date + timedelta(days=1)
    for target_date, label in ((today_date, "Today"), (tomorrow_date, "Tomorrow")):
        try:
            fetch_homepage_slate(target_date, label)
        except Exception:
            continue


def ensure_homepage_cache_refresh_thread() -> None:
    global HOMEPAGE_CACHE_REFRESH_THREAD_STARTED
    with HOMEPAGE_CACHE_REFRESH_THREAD_LOCK:
        if HOMEPAGE_CACHE_REFRESH_THREAD_STARTED:
            return
        HOMEPAGE_CACHE_REFRESH_THREAD_STARTED = True

    def runner() -> None:
        while True:
            refresh_homepage_supporting_caches_once()
            time.sleep(HOMEPAGE_SCOREBOARD_CACHE_SECONDS)

    threading.Thread(target=runner, name="homepage-cache-refresh", daemon=True).start()


def workbook_cell_to_display(value: object) -> object:
    if value in (None, ""):
        return ""
    if isinstance(value, str) and value.strip() == "#N/A":
        return ""
    return value


def starting_pitcher_numeric(value: object) -> Optional[float]:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip().replace(",", "")
        if not stripped:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def derive_starting_pitcher_ip(pitches: object, pitches_per_out: object) -> object:
    pitches_value = starting_pitcher_numeric(pitches)
    pitches_per_out_value = starting_pitcher_numeric(pitches_per_out)
    if pitches_value is None or pitches_per_out_value in (None, 0):
        return ""
    innings_pitched = pitches_value / (pitches_per_out_value * 3)
    return round(innings_pitched, 2)


def build_starting_pitcher_payload() -> Dict[str, object]:
    if not MLB_PRICING_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Could not find workbook at {MLB_PRICING_WORKBOOK_PATH}")

    from openpyxl import load_workbook

    overrides_payload = load_starting_pitcher_overrides_file()
    workbook_mtime = datetime.fromtimestamp(MLB_PRICING_WORKBOOK_PATH.stat().st_mtime).isoformat()
    overrides_mtime = (
        datetime.fromtimestamp(STARTING_PITCHER_OVERRIDES_PATH.stat().st_mtime).isoformat()
        if STARTING_PITCHER_OVERRIDES_PATH.exists()
        else None
    )
    cache_key = json.dumps(
        {
            "workbook_modified_at": workbook_mtime,
            "overrides_modified_at": overrides_mtime,
        },
        sort_keys=True,
    )

    with STARTING_PITCHER_RUNTIME_CACHE_LOCK:
        if STARTING_PITCHER_RUNTIME_CACHE.get("cache_key") == cache_key and STARTING_PITCHER_RUNTIME_CACHE.get("payload"):
            return dict(STARTING_PITCHER_RUNTIME_CACHE["payload"])

    pitcher_overrides = overrides_payload.get("pitchers", {}) if isinstance(overrides_payload, dict) else {}
    rows: List[Dict[str, object]] = []

    workbook = load_workbook(MLB_PRICING_WORKBOOK_PATH, data_only=True, read_only=True)
    try:
        if "SPs" not in workbook.sheetnames:
            raise ValueError("Workbook does not contain an 'SPs' tab.")

        worksheet = workbook["SPs"]
        row_iter = worksheet.iter_rows(values_only=True)
        header_values = next(row_iter, None)
        if not header_values:
            raise ValueError("Workbook 'SPs' tab is empty.")
        headers = [str(value or "").strip() for value in header_values]

        for row_number, raw_values in enumerate(row_iter, start=2):
            row_values = [workbook_cell_to_display(value) for value in raw_values[: len(headers)]]
            if len(row_values) < len(headers):
                row_values.extend([""] * (len(headers) - len(row_values)))

            pitcher_name = str(row_values[0] or "").strip()
            if not pitcher_name:
                continue

            row_data = {headers[index]: row_values[index] for index in range(len(headers))}
            row_data["_row"] = row_number
            row_data["_pitcher"] = pitcher_name
            row_data["_editable"] = {"ERA": True, "Pitches": True}

            overrides = pitcher_overrides.get(pitcher_name, {}) if isinstance(pitcher_overrides, dict) else {}
            if isinstance(overrides, dict):
                for field in ("ERA", "Pitches"):
                    if field in overrides and overrides[field] not in (None, ""):
                        row_data[field] = overrides[field]
                row_data["_override_saved_at"] = overrides.get("saved_at")

            if "IP" in row_data and "P/out" in row_data:
                row_data["IP"] = derive_starting_pitcher_ip(row_data.get("Pitches"), row_data.get("P/out"))

            rows.append(row_data)
    finally:
        workbook.close()

    rows.sort(
        key=lambda row: (
            starting_pitcher_numeric(row.get("ERA")) is None,
            starting_pitcher_numeric(row.get("ERA")) if starting_pitcher_numeric(row.get("ERA")) is not None else 9999.0,
            str(row.get("Pitcher") or ""),
        )
    )
    for index, row in enumerate(rows, start=1):
        row["_rank"] = index

    payload = {
        "headers": headers,
        "rows": rows,
        "source": {
            "workbook_path": str(MLB_PRICING_WORKBOOK_PATH),
            "sheet_name": "SPs",
            "workbook_modified_at": workbook_mtime,
            "overrides_saved_at": overrides_payload.get("saved_at") if isinstance(overrides_payload, dict) else None,
        },
    }
    with STARTING_PITCHER_RUNTIME_CACHE_LOCK:
        STARTING_PITCHER_RUNTIME_CACHE.clear()
        STARTING_PITCHER_RUNTIME_CACHE.update({"cache_key": cache_key, "payload": payload})
    return payload


def build_starting_pitcher_preview_lookup_payload() -> Dict[str, object]:
    workbook_mtime = datetime.fromtimestamp(MLB_PRICING_WORKBOOK_PATH.stat().st_mtime).isoformat()
    overrides_mtime = (
        datetime.fromtimestamp(STARTING_PITCHER_OVERRIDES_PATH.stat().st_mtime).isoformat()
        if STARTING_PITCHER_OVERRIDES_PATH.exists()
        else None
    )
    cache_key = json.dumps(
        {
            "workbook_modified_at": workbook_mtime,
            "overrides_modified_at": overrides_mtime,
        },
        sort_keys=True,
    )

    payload = build_starting_pitcher_payload()
    lookup: Dict[str, Dict[str, object]] = {}
    surname_buckets: Dict[str, List[Dict[str, object]]] = {}

    for row in payload.get("rows", []):
        if not isinstance(row, dict):
            continue
        pitcher_name = str(row.get("Pitcher") or "").strip()
        if not pitcher_name:
            continue
        preview = {
            "pitcher_name": pitcher_name,
            "era": row.get("ERA"),
            "ip": row.get("IP"),
        }
        normalized_full = normalize_pitcher_lookup_name(pitcher_name)
        if normalized_full:
            lookup[normalized_full] = preview

        tokens = re.findall(r"[A-Za-z0-9']+", pitcher_name)
        if tokens:
            surname_key = normalize_pitcher_lookup_name(tokens[-1])
            if surname_key:
                surname_buckets.setdefault(surname_key, []).append(preview)

    for surname_key, previews in surname_buckets.items():
        if len(previews) == 1 and surname_key not in lookup:
            lookup[surname_key] = previews[0]

    return {
        "cache_key": cache_key,
        "generated_at": eastern_now().isoformat(),
        "lookup": lookup,
    }


def get_or_build_starting_pitcher_preview_lookup(force_refresh: bool = False) -> Dict[str, Dict[str, object]]:
    if not MLB_PRICING_WORKBOOK_PATH.exists():
        return {}

    workbook_mtime = datetime.fromtimestamp(MLB_PRICING_WORKBOOK_PATH.stat().st_mtime).isoformat()
    overrides_mtime = (
        datetime.fromtimestamp(STARTING_PITCHER_OVERRIDES_PATH.stat().st_mtime).isoformat()
        if STARTING_PITCHER_OVERRIDES_PATH.exists()
        else None
    )
    cache_key = json.dumps(
        {
            "workbook_modified_at": workbook_mtime,
            "overrides_modified_at": overrides_mtime,
        },
        sort_keys=True,
    )

    if not force_refresh:
        with STARTING_PITCHER_PREVIEW_RUNTIME_CACHE_LOCK:
            if (
                STARTING_PITCHER_PREVIEW_RUNTIME_CACHE.get("cache_key") == cache_key
                and isinstance(STARTING_PITCHER_PREVIEW_RUNTIME_CACHE.get("lookup"), dict)
            ):
                return dict(STARTING_PITCHER_PREVIEW_RUNTIME_CACHE["lookup"])

        cached_file = load_starting_pitcher_preview_cache_file()
        if (
            isinstance(cached_file, dict)
            and cached_file.get("cache_key") == cache_key
            and isinstance(cached_file.get("lookup"), dict)
        ):
            with STARTING_PITCHER_PREVIEW_RUNTIME_CACHE_LOCK:
                STARTING_PITCHER_PREVIEW_RUNTIME_CACHE.clear()
                STARTING_PITCHER_PREVIEW_RUNTIME_CACHE.update(
                    {"cache_key": cache_key, "lookup": cached_file["lookup"]}
                )
            return dict(cached_file["lookup"])

    preview_payload = build_starting_pitcher_preview_lookup_payload()
    save_starting_pitcher_preview_cache_file(preview_payload)
    with STARTING_PITCHER_PREVIEW_RUNTIME_CACHE_LOCK:
        STARTING_PITCHER_PREVIEW_RUNTIME_CACHE.clear()
        STARTING_PITCHER_PREVIEW_RUNTIME_CACHE.update(
            {"cache_key": preview_payload["cache_key"], "lookup": preview_payload["lookup"]}
        )
    return dict(preview_payload["lookup"])


def update_starting_pitcher_override(pitcher_name: str, era: object, pitches: object) -> Dict[str, object]:
    payload = load_starting_pitcher_overrides_file()
    pitcher_payload = payload.setdefault("pitchers", {}) if isinstance(payload, dict) else {}
    if not isinstance(pitcher_payload, dict):
        pitcher_payload = {}
        payload["pitchers"] = pitcher_payload

    def normalize_numeric(value: object) -> object:
        if value in (None, ""):
            return ""
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                return ""
            try:
                numeric_value = float(stripped)
            except ValueError:
                return stripped
            return int(numeric_value) if numeric_value.is_integer() else round(numeric_value, 3)
        if isinstance(value, (int, float)):
            return int(value) if float(value).is_integer() else round(float(value), 3)
        return value

    pitcher_payload[pitcher_name] = {
        "ERA": normalize_numeric(era),
        "Pitches": normalize_numeric(pitches),
        "saved_at": eastern_now().isoformat(),
    }
    payload["saved_at"] = eastern_now().isoformat()
    save_starting_pitcher_overrides_file(payload)
    with STARTING_PITCHER_RUNTIME_CACHE_LOCK:
        STARTING_PITCHER_RUNTIME_CACHE.clear()
    with STARTING_PITCHER_PREVIEW_RUNTIME_CACHE_LOCK:
        STARTING_PITCHER_PREVIEW_RUNTIME_CACHE.clear()
    return pitcher_payload[pitcher_name]


def fetch_depth_chart_data(team_slug: str) -> Dict[str, object]:
    if team_slug in DEPTH_CHART_DATA_CACHE:
        return DEPTH_CHART_DATA_CACHE[team_slug]

    try:
        page_html = fetch_fangraphs_page(f"https://www.fangraphs.com/roster-resource/depth-charts/{team_slug}")
        payload = extract_next_data(page_html)
        queries = payload.get("props", {}).get("pageProps", {}).get("dehydratedState", {}).get("queries", [])
    except Exception:
        DEPTH_CHART_DATA_CACHE[team_slug] = {}
        return {}

    if not queries:
        DEPTH_CHART_DATA_CACHE[team_slug] = {}
        return {}

    data = queries[0].get("state", {}).get("data", {})
    DEPTH_CHART_DATA_CACHE[team_slug] = data if isinstance(data, dict) else {}
    return DEPTH_CHART_DATA_CACHE[team_slug]


def fetch_depth_chart_probables(team_slug: str) -> List[Dict[str, object]]:
    if team_slug in DEPTH_CHART_CACHE:
        return DEPTH_CHART_CACHE[team_slug]

    probable_data = fetch_depth_chart_data(team_slug).get("dataProbableStarters", {})
    game_list = probable_data.get("gameList", []) if isinstance(probable_data, dict) else []
    DEPTH_CHART_CACHE[team_slug] = game_list
    return game_list


def build_fangraphs_pitcher_id_map() -> Dict[str, int]:
    global FANGRAPHS_PITCHER_ID_MAP_CACHE
    if FANGRAPHS_PITCHER_ID_MAP_CACHE is not None:
        return FANGRAPHS_PITCHER_ID_MAP_CACHE

    pitcher_map: Dict[str, int] = {}
    for team in TEAM_OPTIONS:
        try:
            roster_rows = fetch_depth_chart_data(team["slug"]).get("dataRoster", [])
        except Exception:
            roster_rows = []
        if not isinstance(roster_rows, list):
            continue
        for row in roster_rows:
            position_label = str(row.get("position1") or row.get("position") or "")
            roster_type = str(row.get("type") or "").lower()
            mlevel = str(row.get("mlevel") or "").upper()
            mlb_id = row.get("mlbamid")
            fangraphs_id = row.get("playerid") or row.get("playerid1") or row.get("playerid2")
            if "P" not in position_label or not mlb_id or not fangraphs_id:
                continue
            if not (roster_type.startswith("mlb-") or mlevel == "MLB"):
                continue
            pitcher_map[str(mlb_id)] = int(fangraphs_id)

    FANGRAPHS_PITCHER_ID_MAP_CACHE = pitcher_map
    return pitcher_map


def resolve_canonical_fangraphs_pitcher_id(
    mlb_player_id: Optional[int],
    fallback_fangraphs_id: Optional[int] = None,
) -> Optional[int]:
    if mlb_player_id:
        canonical = build_fangraphs_pitcher_id_map().get(str(mlb_player_id))
        if canonical:
            return int(canonical)
    return int(fallback_fangraphs_id) if fallback_fangraphs_id else None


def attach_fangraphs_ids(target_date: date, games: List[Dict[str, object]]) -> None:
    target_label = f"{target_date.month}/{target_date.day}/{target_date.year}"
    for game in games:
        away_slug = resolve_team_slug(str(game.get("away_team", "")))
        if not away_slug:
            continue

        try:
            probable_rows = fetch_depth_chart_probables(away_slug)
        except Exception:
            probable_rows = []
        row = next((item for item in probable_rows if item.get("gameDate") == target_label), None)
        if not row:
            continue

        game["away_pitcher"]["fangraphs_player_id"] = int(row["playerId"]) if row.get("playerId") else None
        game["home_pitcher"]["fangraphs_player_id"] = int(row["oppPlayerId"]) if row.get("oppPlayerId") else None


def extract_next_data(page_html: str) -> Dict[str, object]:
    match = re.search(r'<script id="__NEXT_DATA__" type="application/json">(.*?)</script>', page_html)
    if not match:
        raise ValueError("Could not locate Fangraphs structured page data.")
    return json.loads(match.group(1))


def decode_duckduckgo_href(href: str) -> str:
    decoded_href = html.unescape(href)
    parsed = urlparse(decoded_href)
    if parsed.netloc.endswith("duckduckgo.com") and parsed.path == "/l/":
        uddg = parse_qs(parsed.query).get("uddg", [])
        if uddg:
            return unquote(uddg[0])
    if href.startswith("//"):
        return f"https:{decoded_href}"
    return decoded_href


def resolve_fangraphs_player_url(player_name: str) -> Optional[str]:
    cache_key = normalize_name(player_name)
    if cache_key in FANGRAPHS_URL_CACHE:
        return FANGRAPHS_URL_CACHE[cache_key]

    search_url = f"https://duckduckgo.com/html/?q={quote(player_name + ' site:fangraphs.com/players')}"
    req = Request(search_url, headers=DEFAULT_HEADERS)
    with urlopen(req, timeout=30) as resp:
        search_html = resp.read().decode("utf-8", "ignore")

    candidates = re.findall(r'href="([^"]+)"', search_html)
    for href in candidates:
        decoded = decode_duckduckgo_href(href)
        if "fangraphs.com/players/" not in decoded:
            continue
        player_match = re.search(r"https://www\.fangraphs\.com/players/([^/]+)/([^/?#]+)/", decoded)
        if not player_match:
            continue
        slug, player_id = player_match.groups()
        resolved = f"https://www.fangraphs.com/players/{slug}/{player_id}/stats/pitching"
        FANGRAPHS_URL_CACHE[cache_key] = resolved
        return resolved

    FANGRAPHS_URL_CACHE[cache_key] = None
    return None


def format_decimal(value: Optional[object], digits: int = 2) -> Optional[str]:
    if value is None or value == "":
        return None
    try:
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return str(value)


def format_percentage(value: Optional[object], digits: int = 1) -> Optional[str]:
    if value is None or value == "":
        return None
    try:
        return f"{float(value) * 100:.{digits}f}%"
    except (TypeError, ValueError):
        return str(value)


def find_current_pitching_row(data_common: Dict[str, object]) -> Optional[Dict[str, object]]:
    for row in data_common.get("data", []):
        if row.get("type") == 0 and row.get("AbbLevel") == "MLB":
            return row
    return None


def count_nonempty_pitching_plus_fields(row: Optional[Dict[str, object]]) -> int:
    if not isinstance(row, dict):
        return 0
    fields = ("sp_stuff", "sp_location", "sp_pitching")
    return sum(1 for field in fields if row.get(field) not in (None, ""))


def strip_tags(value: object) -> str:
    return re.sub(r"<[^>]+>", "", str(value or "")).strip()


def extract_season_int(value: object) -> Optional[int]:
    cleaned = strip_tags(value)
    match = re.search(r"\b(20\d{2})\b", cleaned)
    return int(match.group(1)) if match else None


def find_pitching_row_for_season(rows: List[Dict[str, object]], season: int) -> Optional[Dict[str, object]]:
    candidates = [
        row
        for row in rows
        if row.get("AbbLevel") == "MLB" and extract_season_int(row.get("Season")) == season
    ]
    if not candidates:
        return None

    def sort_key(row: Dict[str, object]) -> tuple[int, int]:
        team_label = strip_tags(row.get("Team"))
        is_total = 0 if team_label in {"- - -", "TOT"} else 1
        has_pitching_plus = 0 if count_nonempty_pitching_plus_fields(row) > 0 else 1
        type_value = 0 if row.get("type") == 0 else 1
        return (has_pitching_plus, is_total, type_value)

    return sorted(candidates, key=sort_key)[0]


def collect_pitching_rows(page_props: Dict[str, object]) -> List[Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for section_key, data_key in (("dataCommon", "data"), ("dataStats", "data")):
        section = page_props.get(section_key, {})
        section_rows = section.get(data_key, []) if isinstance(section, dict) else []
        if not isinstance(section_rows, list):
            continue
        for row in section_rows:
            if isinstance(row, dict) and row.get("AbbLevel") == "MLB":
                rows.append(row)
    return rows


def find_projection_row(rows: List[Dict[str, object]], team_label: str) -> Optional[Dict[str, object]]:
    for row in rows:
        if str(row.get("Team")) == team_label:
            return row
    return None


def build_pitcher_stuff_comparison_payload(
    *,
    mlb_player_id: int,
    fangraphs_player_id: Optional[int],
    pitcher_name: str,
    fangraphs_url: Optional[str],
    current_year: int,
    previous_year: int,
    current_stuff_value: Optional[object],
    previous_stuff_value: Optional[object],
    stuff_diff_value: Optional[float],
    location_plus_value: Optional[object],
    pitching_plus_value: Optional[object],
) -> Dict[str, object]:
    canonical_fangraphs_id = resolve_canonical_fangraphs_pitcher_id(mlb_player_id, fangraphs_player_id)
    canonical_fangraphs_url = (
        f"https://www.fangraphs.com/statss.aspx?playerid={canonical_fangraphs_id}"
        if canonical_fangraphs_id
        else fangraphs_url
    )
    return {
        "cache_saved_at": eastern_now().isoformat(),
        "mlb_player_id": mlb_player_id,
        "fangraphs_player_id": canonical_fangraphs_id,
        "pitcher_name": pitcher_name,
        "normalized_name": normalize_name(pitcher_name),
        "fangraphs_url": canonical_fangraphs_url,
        "current_season": current_year,
        "previous_season": previous_year,
        "metrics": {
            "StuffPlusCurrent": format_decimal(current_stuff_value, 1),
            "StuffPlusPrevious": format_decimal(previous_stuff_value, 1),
            "StuffPlusDiff": format_decimal(stuff_diff_value, 1),
            "LocationPlus": format_decimal(location_plus_value, 1),
            "PitchingPlus": format_decimal(pitching_plus_value, 1),
        },
    }


def pitcher_stuff_metric_fields() -> List[str]:
    return [
        "StuffPlusCurrent",
        "StuffPlusPrevious",
        "StuffPlusDiff",
        "LocationPlus",
        "PitchingPlus",
    ]


def count_nonempty_pitcher_stuff_metrics(metrics: Dict[str, object]) -> int:
    return sum(1 for field in pitcher_stuff_metric_fields() if metrics.get(field) not in (None, ""))


def pitcher_stuff_entry_has_useful_values(entry: Optional[Dict[str, object]], current_year: int, previous_year: int) -> bool:
    if not entry:
        return False
    if entry.get("current_season") != current_year or entry.get("previous_season") != previous_year:
        return False
    metrics = entry.get("metrics", {}) if isinstance(entry.get("metrics"), dict) else {}
    return count_nonempty_pitcher_stuff_metrics(metrics) >= 3 and metrics.get("StuffPlusCurrent") not in (None, "")


def pitcher_stuff_entry_needs_refresh(entry: Optional[Dict[str, object]], current_year: int, previous_year: int) -> bool:
    if not entry:
        return True
    if entry.get("current_season") != current_year or entry.get("previous_season") != previous_year:
        return True
    metrics = entry.get("metrics", {}) if isinstance(entry.get("metrics"), dict) else {}
    if metrics.get("StuffPlusCurrent") in (None, ""):
        return True
    if metrics.get("StuffPlusPrevious") in (None, ""):
        return True
    return False


def merge_pitcher_stuff_comparison_entries(
    existing: Optional[Dict[str, object]],
    incoming: Dict[str, object],
) -> Dict[str, object]:
    if not existing:
        return incoming

    existing_metrics = existing.get("metrics", {}) if isinstance(existing.get("metrics"), dict) else {}
    incoming_metrics = incoming.get("metrics", {}) if isinstance(incoming.get("metrics"), dict) else {}
    merged_metrics: Dict[str, object] = {}
    for field in pitcher_stuff_metric_fields():
        incoming_value = incoming_metrics.get(field)
        existing_value = existing_metrics.get(field)
        merged_metrics[field] = incoming_value if incoming_value not in (None, "") else existing_value

    current_value = merged_metrics.get("StuffPlusCurrent")
    previous_value = merged_metrics.get("StuffPlusPrevious")
    try:
        if current_value not in (None, "") and previous_value not in (None, ""):
            merged_metrics["StuffPlusDiff"] = format_decimal(float(current_value) - float(previous_value), 1)
        elif merged_metrics.get("StuffPlusDiff") in (None, ""):
            merged_metrics["StuffPlusDiff"] = existing_metrics.get("StuffPlusDiff")
    except (TypeError, ValueError):
        pass

    existing_count = count_nonempty_pitcher_stuff_metrics(existing_metrics)
    incoming_count = count_nonempty_pitcher_stuff_metrics(incoming_metrics)
    saved_at = incoming.get("cache_saved_at") if incoming_count >= existing_count else existing.get("cache_saved_at")

    canonical_fangraphs_id = incoming.get("fangraphs_player_id") or existing.get("fangraphs_player_id")
    canonical_fangraphs_url = (
        f"https://www.fangraphs.com/statss.aspx?playerid={canonical_fangraphs_id}"
        if canonical_fangraphs_id
        else incoming.get("fangraphs_url") or existing.get("fangraphs_url")
    )

    return {
        **existing,
        **incoming,
        "cache_saved_at": saved_at,
        "mlb_player_id": incoming.get("mlb_player_id") or existing.get("mlb_player_id"),
        "fangraphs_player_id": canonical_fangraphs_id,
        "fangraphs_url": canonical_fangraphs_url,
        "pitcher_name": incoming.get("pitcher_name") or existing.get("pitcher_name"),
        "normalized_name": incoming.get("normalized_name") or existing.get("normalized_name"),
        "current_season": incoming.get("current_season") or existing.get("current_season"),
        "previous_season": incoming.get("previous_season") or existing.get("previous_season"),
        "metrics": merged_metrics,
    }


def build_pitcher_profile_from_cached_stuff(
    *,
    mlb_player_id: int,
    person: Dict[str, object],
    full_name: str,
    fangraphs_url: Optional[str],
    cached_stuff: Dict[str, object],
) -> Dict[str, object]:
    metrics = cached_stuff.get("metrics", {}) if isinstance(cached_stuff, dict) else {}
    current_year = cached_stuff.get("current_season", eastern_now().year)
    previous_year = cached_stuff.get("previous_season", int(current_year) - 1 if str(current_year).isdigit() else eastern_now().year - 1)
    return {
        "player": {
            "mlb_id": mlb_player_id,
            "name": person.get("fullName") or full_name,
            "team": person.get("currentTeam", {}).get("name") if isinstance(person.get("currentTeam"), dict) else None,
            "number": person.get("primaryNumber"),
            "headshot_url": MLB_HEADSHOT_TEMPLATE.format(mlb_id=mlb_player_id),
            "height": person.get("height"),
            "weight": person.get("weight"),
            "age": person.get("currentAge"),
            "pitch_hand": (person.get("pitchHand") or {}).get("description", ""),
        },
        "sources": {
            "mlb_url": f"https://www.mlb.com/player/{person.get('nameSlug') or mlb_player_id}",
            "fangraphs_url": fangraphs_url or cached_stuff.get("fangraphs_url") or "",
        },
        "seasons": {
            "current": current_year,
            "previous": previous_year,
        },
        "cache": {
            "stuff_comparison_saved_at": cached_stuff.get("cache_saved_at"),
            "stuff_comparison_is_cached_fallback": True,
        },
        "sp_board": build_starting_pitcher_profile_section(person.get("fullName") or full_name),
        "metrics": {
            "ERA": None,
            "xERA": None,
            "FIP": None,
            "xFIP": None,
            "SIERA": None,
            "StuffPlus": metrics.get("StuffPlusCurrent"),
            "StuffPlusPrevious": metrics.get("StuffPlusPrevious"),
            "StuffPlusDiff": metrics.get("StuffPlusDiff"),
            "LocationPlus": metrics.get("LocationPlus"),
            "PitchingPlus": metrics.get("PitchingPlus"),
            "SteamerRosERA": None,
            "TheBatRosERA": None,
            "KMinusBBPercent": None,
        },
    }


def load_pitcher_profile_payload_cache_file() -> Dict[str, object]:
    if not PITCHER_PROFILE_PAYLOAD_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(PITCHER_PROFILE_PAYLOAD_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_pitcher_profile_payload_cache_file(payload: Dict[str, object]) -> None:
    PITCHER_PROFILE_PAYLOAD_CACHE_PATH.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def build_pitcher_profile_cache_key(mlb_player_id: int, fangraphs_player_id: Optional[int]) -> str:
    workbook_mtime = MLB_PRICING_WORKBOOK_PATH.stat().st_mtime if MLB_PRICING_WORKBOOK_PATH.exists() else 0
    overrides_mtime = STARTING_PITCHER_OVERRIDES_PATH.stat().st_mtime if STARTING_PITCHER_OVERRIDES_PATH.exists() else 0
    stuff_cache_mtime = (
        PITCHER_STUFF_COMPARISON_CACHE_PATH.stat().st_mtime
        if PITCHER_STUFF_COMPARISON_CACHE_PATH.exists()
        else 0
    )
    return json.dumps(
        {
            "mlb_player_id": mlb_player_id,
            "fangraphs_player_id": fangraphs_player_id or 0,
            "season": eastern_now().year,
            "workbook_mtime": workbook_mtime,
            "overrides_mtime": overrides_mtime,
            "stuff_cache_mtime": stuff_cache_mtime,
        },
        sort_keys=True,
    )


def get_cached_pitcher_profile_payload(
    mlb_player_id: int,
    cache_key: str,
) -> Optional[Dict[str, object]]:
    def payload_is_usable(payload: object) -> bool:
        if not isinstance(payload, dict):
            return False
        player = payload.get("player", {})
        return isinstance(player, dict) and bool(player.get("team"))

    runtime_entry = PITCHER_PROFILE_CACHE.get(mlb_player_id)
    if isinstance(runtime_entry, dict):
        if (
            runtime_entry.get("_cache_key") == cache_key
            and cache_timestamp_is_fresh(runtime_entry.get("_saved_at"), PITCHER_PROFILE_CACHE_SECONDS)
            and payload_is_usable(runtime_entry.get("payload"))
        ):
            return dict(runtime_entry["payload"])

    cached = load_pitcher_profile_payload_cache_file()
    entries = cached.get("pitchers", {}) if isinstance(cached, dict) else {}
    entry = entries.get(str(mlb_player_id)) if isinstance(entries, dict) else None
    if not isinstance(entry, dict):
        return None
    if entry.get("cache_key") != cache_key:
        return None
    if not cache_timestamp_is_fresh(entry.get("saved_at"), PITCHER_PROFILE_CACHE_SECONDS):
        return None
    payload = entry.get("payload")
    if not payload_is_usable(payload):
        return None
    PITCHER_PROFILE_CACHE[mlb_player_id] = {
        "_cache_key": cache_key,
        "_saved_at": entry.get("saved_at"),
        "payload": payload,
    }
    return dict(payload)


def save_cached_pitcher_profile_payload(
    mlb_player_id: int,
    cache_key: str,
    payload: Dict[str, object],
) -> None:
    saved_at = eastern_now().isoformat()
    PITCHER_PROFILE_CACHE[mlb_player_id] = {
        "_cache_key": cache_key,
        "_saved_at": saved_at,
        "payload": payload,
    }
    cached = load_pitcher_profile_payload_cache_file()
    entries = cached.setdefault("pitchers", {}) if isinstance(cached, dict) else {}
    if not isinstance(entries, dict):
        entries = {}
        cached["pitchers"] = entries
    entries[str(mlb_player_id)] = {
        "cache_key": cache_key,
        "saved_at": saved_at,
        "payload": payload,
    }
    cached["saved_at"] = saved_at
    save_pitcher_profile_payload_cache_file(cached)


def build_pitcher_profile(mlb_player_id: int, fangraphs_player_id: Optional[int] = None) -> Dict[str, object]:
    canonical_fangraphs_player_id = resolve_canonical_fangraphs_pitcher_id(mlb_player_id, fangraphs_player_id)
    cache_key = build_pitcher_profile_cache_key(mlb_player_id, canonical_fangraphs_player_id)
    cached_profile = get_cached_pitcher_profile_payload(mlb_player_id, cache_key)
    if cached_profile:
        return cached_profile

    person = fetch_mlb_person(mlb_player_id)
    full_name = person.get("fullName", "")
    cached_stuff = get_cached_pitcher_stuff_comparison(mlb_player_id, canonical_fangraphs_player_id)
    fangraphs_url = (
        cached_stuff.get("fangraphs_url")
        if isinstance(cached_stuff, dict) and cached_stuff.get("fangraphs_url")
        else (
            f"https://www.fangraphs.com/statss.aspx?playerid={canonical_fangraphs_player_id}"
            if canonical_fangraphs_player_id
            else None
        )
    )
    if cached_stuff:
        profile = build_pitcher_profile_from_cached_stuff(
            mlb_player_id=mlb_player_id,
            person=person,
            full_name=full_name,
            fangraphs_url=fangraphs_url,
            cached_stuff=cached_stuff,
        )
        save_cached_pitcher_profile_payload(mlb_player_id, cache_key, profile)
        return profile

    if not fangraphs_url:
        fangraphs_url = resolve_fangraphs_player_url(full_name)
    if not fangraphs_url:
        raise ValueError("Could not resolve a Fangraphs player page for this pitcher.")

    try:
        fg_html = fetch_fangraphs_page(fangraphs_url)
    except Exception:
        if cached_stuff:
            profile = build_pitcher_profile_from_cached_stuff(
                mlb_player_id=mlb_player_id,
                person=person,
                full_name=full_name,
                fangraphs_url=fangraphs_url,
                cached_stuff=cached_stuff,
            )
            save_cached_pitcher_profile_payload(mlb_player_id, cache_key, profile)
            return profile
        raise

    fg_payload = extract_next_data(fg_html)
    page_props = fg_payload.get("props", {}).get("pageProps", {})
    data_common = page_props.get("dataCommon", {})
    data_stats = page_props.get("dataStats", {})
    player_info = data_stats.get("playerInfo", {})
    team_info = data_stats.get("teamInfo", {})
    current_year = eastern_now().year
    previous_year = current_year - 1
    season_rows = data_common.get("data", [])
    current_row = find_current_pitching_row(data_common)
    current_season_row = find_pitching_row_for_season(season_rows, current_year) or current_row
    previous_season_row = find_pitching_row_for_season(season_rows, previous_year)
    projection_rows = data_stats.get("data", [])

    if not current_row:
        raise ValueError("Could not find the current Fangraphs pitching row for this pitcher.")

    steamer_ros = find_projection_row(projection_rows, "Steamer (RoS)")
    the_bat_ros = find_projection_row(projection_rows, "THE BAT (RoS)")
    comparison_payload = fetch_pitcher_stuff_comparison_entry(
        mlb_player_id=mlb_player_id,
        pitcher_name=person.get("fullName") or player_info.get("firstLastName") or full_name,
        fangraphs_player_id=canonical_fangraphs_player_id,
        current_year=current_year,
        previous_year=previous_year,
    )
    comparison_metrics = comparison_payload.get("metrics", {})
    current_stuff_value = comparison_metrics.get("StuffPlusCurrent")
    previous_stuff_value = comparison_metrics.get("StuffPlusPrevious")
    stuff_diff_value = comparison_metrics.get("StuffPlusDiff")
    location_plus_value = comparison_metrics.get("LocationPlus")
    pitching_plus_value = comparison_metrics.get("PitchingPlus")

    save_cached_pitcher_stuff_comparison(comparison_payload)

    profile = {
        "player": {
            "mlb_id": mlb_player_id,
            "name": person.get("fullName") or player_info.get("firstLastName") or full_name,
            "team": team_info.get("MLB_FullName"),
            "number": person.get("primaryNumber"),
            "headshot_url": MLB_HEADSHOT_TEMPLATE.format(mlb_id=mlb_player_id),
            "height": person.get("height") or player_info.get("HeightDisplay"),
            "weight": person.get("weight") or player_info.get("Weight"),
            "age": person.get("currentAge") or player_info.get("AgeToday"),
            "pitch_hand": (person.get("pitchHand") or {}).get("description", ""),
        },
        "sources": {
            "mlb_url": f"https://www.mlb.com/player/{person.get('nameSlug') or mlb_player_id}",
            "fangraphs_url": fangraphs_url,
        },
        "seasons": {
            "current": current_year,
            "previous": previous_year,
        },
        "cache": {
            "stuff_comparison_saved_at": eastern_now().isoformat(),
            "stuff_comparison_is_cached_fallback": False,
        },
        "sp_board": build_starting_pitcher_profile_section(person.get("fullName") or player_info.get("firstLastName") or full_name),
        "metrics": {
            "ERA": format_decimal(current_row.get("ERA")),
            "xERA": format_decimal(current_row.get("xERA")),
            "FIP": format_decimal(current_row.get("FIP")),
            "xFIP": format_decimal(current_row.get("xFIP")),
            "SIERA": format_decimal(current_row.get("SIERA")),
            "StuffPlus": current_stuff_value,
            "StuffPlusPrevious": previous_stuff_value,
            "StuffPlusDiff": stuff_diff_value,
            "LocationPlus": location_plus_value,
            "PitchingPlus": pitching_plus_value,
            "SteamerRosERA": format_decimal((steamer_ros or {}).get("ERA")),
            "TheBatRosERA": format_decimal((the_bat_ros or {}).get("ERA")),
            "KMinusBBPercent": format_percentage(current_row.get("K-BB%")),
        },
    }

    save_cached_pitcher_profile_payload(mlb_player_id, cache_key, profile)
    return profile


def find_section_indexes(lines: List[str]) -> Dict[str, int]:
    indexes: Dict[str, int] = {}
    for idx, line in enumerate(lines):
        compact = line.lower()
        if "starting lineup vsr" in compact and "vsR" not in indexes:
            indexes["vsR"] = idx
        elif "starting lineup vsl" in compact and "vsL" not in indexes:
            indexes["vsL"] = idx
        elif line.startswith("Updated:") and "updated" not in indexes:
            indexes["updated"] = idx
    return indexes


def parse_header_row(lines: List[str], start: int) -> List[str]:
    for idx in range(start, min(start + 4, len(lines))):
        if lines[idx].startswith("Order/Status"):
            header = lines[idx].replace("Order/Status", "Order/Status ")
            header = header.replace("wRC+", " wRC+")
            parts = [part for part in header.split() if part]
            return parts
    return ["Order/Status", "Position", "Name", "Bats", "PA", "HR", "wRC+"]


def locate_bats_index(tokens: List[str], position_index: int) -> Optional[int]:
    for idx in range(len(tokens) - 1, position_index, -1):
        if tokens[idx] in VALID_BATS:
            return idx
    return None


def parse_player_row(line: str) -> Optional[PlayerRow]:
    if line.startswith("Order/Status") or line.startswith("Updated:"):
        return None

    cleaned = re.sub(r"[▼▲â–¼â–²]", " ", line)
    cleaned = cleaned.replace("Ã¢â€“Â¼", " ").replace("Ã¢â€“Â²", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    tokens = cleaned.split()
    if len(tokens) < 4:
        return None

    position_index = next((idx for idx, token in enumerate(tokens) if token in POSITIONS), None)
    if position_index is None or position_index == 0:
        return None

    bats_index = locate_bats_index(tokens, position_index)
    if bats_index is None or bats_index <= position_index + 1:
        return None

    status = " ".join(tokens[:position_index])
    position = tokens[position_index]
    name = " ".join(tokens[position_index + 1:bats_index])
    stats = tokens[bats_index + 1:]

    if not status or not position or not name:
        return None

    return PlayerRow(status=status, position=position, name=name, bats=tokens[bats_index], stats=stats)


def parse_section(lines: List[str], start: int, end: int) -> Dict[str, object]:
    title = lines[start]
    headers = parse_header_row(lines, start + 1)
    rows: List[PlayerRow] = []
    seen = set()

    for line in lines[start + 1:end]:
        row = parse_player_row(line)
        if row is None:
            continue
        row_key = (row.status, row.position, row.name, row.bats, tuple(row.stats))
        if row_key in seen:
            continue
        seen.add(row_key)
        rows.append(row)

    return {
        "title": title,
        "headers": headers,
        "rows": [
            {
                "status": row.status,
                "position": row.position,
                "name": row.name,
                "bats": row.bats,
                "stats": row.stats,
            }
            for row in rows
        ],
    }


def split_rows(rows: List[Dict[str, object]]) -> Dict[str, List[str]]:
    starters: List[str] = []
    bench: List[str] = []
    injured: List[str] = []

    for row in rows:
        status = str(row.get("status", "")).strip()
        name = str(row.get("name", "")).strip()
        if not name:
            continue
        lowered = status.lower()
        if lowered.startswith("bench"):
            bench.append(name)
        elif lowered.startswith("il"):
            injured.append(name)
        elif re.match(r"^\d+$", status):
            starters.append(name)

    if not starters:
        starters = [str(row.get("name", "")).strip() for row in rows[:STARTING_LINEUP_LENGTH] if row.get("name")]

    return {"starters": starters, "bench": bench, "il": injured}


def is_fangraphs_interstitial(raw_html: str) -> bool:
    lowered = (raw_html or "").lower()
    challenge_markers = (
        "<title>just a moment...</title>",
        "cf-mitigated",
        "cdn-cgi/challenge-platform",
        "challenges.cloudflare.com",
        "enable javascript and cookies to continue",
    )
    return any(marker in lowered for marker in challenge_markers)


def parse_lineup_page(team_slug: str, raw_html: str) -> Dict[str, object]:
    if team_slug not in TEAM_MAP:
        raise HTTPException(status_code=404, detail="Unknown team slug")

    if is_fangraphs_interstitial(raw_html):
        raise ValueError("Fangraphs returned a Cloudflare challenge page instead of lineup data.")

    lines = extract_lines(raw_html)
    indexes = find_section_indexes(lines)
    if "vsR" not in indexes or "vsL" not in indexes:
        raise ValueError("Could not find Fangraphs platoon lineup sections on the page.")

    updated_line = lines[indexes["updated"]] if "updated" in indexes else "Updated: unavailable"
    vsr_end = indexes["vsL"]
    vsl_end = indexes.get("updated", len(lines))
    vsr_section = parse_section(lines, indexes["vsR"], vsr_end)
    vsl_section = parse_section(lines, indexes["vsL"], vsl_end)
    if not vsr_section["rows"] and not vsl_section["rows"]:
        raise ValueError("Parsed the Fangraphs page but did not find any lineup rows.")

    return {
        "team": TEAM_MAP[team_slug],
        "source_url": f"{BASE_URL}/{team_slug}",
        "updated": updated_line,
        "refreshed_every_seconds": REFRESH_SECONDS,
        "sections": {
            "vsR": vsr_section,
            "vsL": vsl_section,
        },
        "excel_layout": {
            "vsR": split_rows(vsr_section["rows"]),
            "vsL": split_rows(vsl_section["rows"]),
        },
    }


def get_cached_lineup_data(team_slug: str) -> Optional[Dict[str, object]]:
    cached = load_lineup_cache_file()
    payload = cached.get("teams", {}).get(team_slug) if isinstance(cached, dict) else None
    return payload if isinstance(payload, dict) else None


def save_cached_lineup_data(team_slug: str, payload: Dict[str, object]) -> None:
    cached = load_lineup_cache_file()
    teams_payload = cached.setdefault("teams", {}) if isinstance(cached, dict) else {}
    if not isinstance(teams_payload, dict):
        teams_payload = {}
        cached["teams"] = teams_payload
    teams_payload[team_slug] = {
        **payload,
        "cache_saved_at": eastern_now().isoformat(),
    }
    save_lineup_cache_file(cached)


def fetch_lineup_data(team_slug: str) -> Dict[str, object]:
    url = f"{BASE_URL}/{team_slug}"
    last_error: Optional[Exception] = None

    for delay_seconds in (0.0, 1.0):
        if delay_seconds:
            time.sleep(delay_seconds)
        try:
            html_text = fetch_fangraphs_page(url)
            payload = parse_lineup_page(team_slug, html_text)
            save_cached_lineup_data(team_slug, payload)
            return payload
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            last_error = exc

    cached_payload = get_cached_lineup_data(team_slug)
    if cached_payload:
        return cached_payload

    if last_error:
        raise last_error
    raise ValueError("Failed to fetch Fangraphs lineup page and no cached lineup data is available.")


TEAM_CODE_MAP = {team["code"]: team for team in TEAM_OPTIONS}
BULLPEN_TEAM_CODE_ALIASES = {
    "ATH": "athletics",
}
BULLPEN_ROLE_PREFIXES = [
    "Co-Closer",
    "Closer Committee",
    "Middle Reliever",
    "Long Reliever",
    "Setup Man",
    "Closer",
]


def strip_fangraphs_link_markup(line: str) -> str:
    cleaned = re.sub(r"【\d+†", " ", line)
    cleaned = cleaned.replace("】", " ")
    return re.sub(r"\s+", " ", cleaned).strip()


def merge_split_bullpen_lines(lines: List[str]) -> List[str]:
    merged: List[str] = []
    continuation_pattern = re.compile(
        r"^[RL]\s+(?:%s)\b" % "|".join(re.escape(role) for role in BULLPEN_ROLE_PREFIXES)
    )

    idx = 0
    while idx < len(lines):
        line = strip_fangraphs_link_markup(lines[idx])
        if (
            idx + 1 < len(lines)
            and re.match(r"^[A-Z]{3}\s+.+$", line)
            and not parse_bullpen_row(line)
        ):
            next_line = strip_fangraphs_link_markup(lines[idx + 1])
            if continuation_pattern.match(next_line):
                merged.append(f"{line} {next_line}")
                idx += 2
                continue

        merged.append(line)
        idx += 1

    return merged


def parse_bullpen_row(line: str) -> Optional[Dict[str, object]]:
    cleaned = strip_fangraphs_link_markup(line)
    match = re.match(r"^(?P<team>[A-Z]{3})\s+(?P<body>.+)$", cleaned)
    if not match:
        return None

    team_code = match.group("team")
    team = TEAM_CODE_MAP.get(team_code)
    if not team:
        team_slug = BULLPEN_TEAM_CODE_ALIASES.get(team_code)
        team = TEAM_MAP.get(team_slug) if team_slug else None
    if not team:
        return None

    body = match.group("body")
    body_match = re.match(r"^(?P<name>.+?)\s+(?P<throws>[RL])\s+(?P<rest>.+)$", body)
    if not body_match:
        return None

    name = body_match.group("name").strip()
    throws = body_match.group("throws")
    rest = body_match.group("rest").strip()

    if rest.endswith("IL"):
        return {
            "kind": "injured",
            "team_code": team_code,
            "row": BullpenInjuryRow(name=name, throws=throws, status=rest),
        }

    role = next((prefix for prefix in BULLPEN_ROLE_PREFIXES if rest.startswith(prefix)), "")
    if not role:
        return None

    remainder = rest[len(role) :].strip()
    tokens = remainder.split()
    if len(tokens) < 9:
        return None

    usage = tokens[:-9]
    stats = tokens[-9:]
    return {
        "kind": "active",
        "team_code": team["code"],
        "row": BullpenRow(name=name, throws=throws, role=role, usage=usage, stats=stats),
    }


def parse_bullpen_page(raw_html: str) -> Dict[str, object]:
    lines = merge_split_bullpen_lines(extract_lines(raw_html))
    updated = next((line for line in reversed(lines) if line.startswith("Updated:")), "Updated: unavailable")
    teams: Dict[str, Dict[str, object]] = {
        team["slug"]: {
            "team": team,
            "active_headers": ["Name", "Thr", "Role", "Recent Usage", "vFA", "vSI", "G", "IP", "ERA", "Sv", "HLD", "K/9", "K%"],
            "active_rows": [],
            "injured_headers": ["Name", "Thr", "Status"],
            "injured_rows": [],
            "active_seen": set(),
            "injured_seen": set(),
        }
        for team in TEAM_OPTIONS
    }

    for line in lines:
        parsed = parse_bullpen_row(line)
        if not parsed:
            continue
        team = TEAM_CODE_MAP[parsed["team_code"]]
        bucket = teams[team["slug"]]
        if parsed["kind"] == "active":
            row = parsed["row"]
            row_key = (row.name, row.throws, row.role, tuple(row.usage), tuple(row.stats))
            if row_key in bucket["active_seen"]:
                continue
            bucket["active_seen"].add(row_key)
            bucket["active_rows"].append(
                {
                    "name": row.name,
                    "throws": row.throws,
                    "role": row.role,
                    "usage": " ".join(row.usage),
                    "stats": row.stats,
                }
            )
        else:
            row = parsed["row"]
            row_key = (row.name, row.throws, row.status)
            if row_key in bucket["injured_seen"]:
                continue
            bucket["injured_seen"].add(row_key)
            bucket["injured_rows"].append(
                {
                    "name": row.name,
                    "throws": row.throws,
                    "status": row.status,
                }
            )

    for team_payload in teams.values():
        team_payload.pop("active_seen", None)
        team_payload.pop("injured_seen", None)

    return {
        "source_url": BULLPEN_URL,
        "updated": updated,
        "refreshed_every_seconds": REFRESH_SECONDS,
        "teams": teams,
    }


def fetch_bullpen_data(team_slug: str) -> Dict[str, object]:
    if team_slug not in TEAM_MAP:
        raise HTTPException(status_code=404, detail="Unknown team slug")

    last_error: Optional[Exception] = None
    payload: Optional[Dict[str, object]] = None

    for delay_seconds in (0.0, 1.0):
        if delay_seconds:
            time.sleep(delay_seconds)
        try:
            raw_html = fetch_fangraphs_page(BULLPEN_URL)
            payload = parse_bullpen_page(raw_html)
            payload = save_cached_bullpen_data(payload)
            break
        except (HTTPError, URLError, TimeoutError, ValueError) as exc:
            last_error = exc

    if payload is None:
        cached_payload = load_bullpen_cache_file()
        if cached_payload:
            payload = cached_payload
        elif last_error:
            raise last_error
        else:
            raise ValueError("Failed to fetch Fangraphs bullpen page and no cached bullpen data is available.")

    team_payload = payload["teams"][team_slug]
    return {
        "team": team_payload["team"],
        "source_url": payload["source_url"],
        "updated": payload["updated"],
        "refreshed_every_seconds": payload["refreshed_every_seconds"],
        "cache_saved_at": payload.get("cache_saved_at"),
        "sections": {
            "active": {
                "title": "Projected Bullpen",
                "headers": team_payload["active_headers"],
                "rows": team_payload["active_rows"],
            },
            "injured": {
                "title": "IL / Unavailable",
                "headers": team_payload["injured_headers"],
                "rows": team_payload["injured_rows"],
            },
        },
    }


def extract_js_array(raw_html: str, marker: str) -> str:
    marker_index = raw_html.find(marker)
    if marker_index == -1:
        raise ValueError("Could not find embedded Baseball Savant pitch arsenal data.")

    array_start = raw_html.find("[", marker_index)
    if array_start == -1:
        raise ValueError("Could not find start of Baseball Savant pitch arsenal data array.")

    depth = 0
    in_string = False
    escape = False

    for idx in range(array_start, len(raw_html)):
        char = raw_html[idx]

        if in_string:
            if escape:
                escape = False
            elif char == "\\":
                escape = True
            elif char == '"':
                in_string = False
            continue

        if char == '"':
            in_string = True
        elif char == "[":
            depth += 1
        elif char == "]":
            depth -= 1
            if depth == 0:
                return raw_html[array_start:idx + 1]

    raise ValueError("Could not find end of Baseball Savant pitch arsenal data array.")


def extract_savant_pitch_data(raw_html: str) -> List[Dict[str, object]]:
    return json.loads(extract_js_array(raw_html, "var data ="))


def eastern_now() -> datetime:
    try:
        return datetime.now(ZoneInfo("America/New_York"))
    except ZoneInfoNotFoundError:
        # On some Windows Python installs tzdata is missing; this app runs on the user's ET machine.
        return datetime.now()


def current_stuff_plus_refresh_key(now: Optional[datetime] = None) -> str:
    current_time = now if now else eastern_now()
    if current_time.tzinfo is not None:
        try:
            current_time = current_time.astimezone(ZoneInfo("America/New_York"))
        except ZoneInfoNotFoundError:
            pass
    effective_date = current_time.date() if current_time.hour >= 4 else current_time.date() - timedelta(days=1)
    return effective_date.isoformat()


def load_stuff_plus_cache_file() -> Dict[str, object]:
    if not STUFF_PLUS_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(STUFF_PLUS_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_stuff_plus_cache_file(payload: Dict[str, object]) -> None:
    STUFF_PLUS_CACHE_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def load_velocity_comparison_cache_file() -> Dict[str, object]:
    if not VELOCITY_COMPARISON_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(VELOCITY_COMPARISON_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_velocity_comparison_cache_file(payload: Dict[str, object]) -> None:
    VELOCITY_COMPARISON_CACHE_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def load_lineup_cache_file() -> Dict[str, object]:
    if not LINEUP_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(LINEUP_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_lineup_cache_file(payload: Dict[str, object]) -> None:
    LINEUP_CACHE_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def load_bullpen_cache_file() -> Dict[str, object]:
    if not BULLPEN_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(BULLPEN_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_bullpen_cache_file(payload: Dict[str, object]) -> None:
    BULLPEN_CACHE_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def load_homepage_scoreboard_cache_file() -> Dict[str, object]:
    if not HOMEPAGE_SCOREBOARD_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(HOMEPAGE_SCOREBOARD_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_homepage_scoreboard_cache_file(payload: Dict[str, object]) -> None:
    HOMEPAGE_SCOREBOARD_CACHE_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def load_starting_pitcher_preview_cache_file() -> Dict[str, object]:
    if not STARTING_PITCHER_PREVIEW_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(STARTING_PITCHER_PREVIEW_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_starting_pitcher_preview_cache_file(payload: Dict[str, object]) -> None:
    STARTING_PITCHER_PREVIEW_CACHE_PATH.write_text(
        json.dumps(payload, indent=2, sort_keys=True),
        encoding="utf-8",
    )


def load_starting_pitcher_overrides_file() -> Dict[str, object]:
    if not STARTING_PITCHER_OVERRIDES_PATH.exists():
        return {}
    try:
        return json.loads(STARTING_PITCHER_OVERRIDES_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_starting_pitcher_overrides_file(payload: Dict[str, object]) -> None:
    STARTING_PITCHER_OVERRIDES_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def save_cached_bullpen_data(payload: Dict[str, object]) -> Dict[str, object]:
    cached_payload = {
        **payload,
        "cache_saved_at": eastern_now().isoformat(),
    }
    save_bullpen_cache_file(cached_payload)
    return cached_payload


def load_pitcher_stuff_comparison_cache_file() -> Dict[str, object]:
    if not PITCHER_STUFF_COMPARISON_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(PITCHER_STUFF_COMPARISON_CACHE_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def save_pitcher_stuff_comparison_cache_file(payload: Dict[str, object]) -> None:
    PITCHER_STUFF_COMPARISON_CACHE_PATH.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")


def get_cached_pitcher_stuff_comparison(
    mlb_player_id: int,
    fangraphs_player_id: Optional[int] = None,
) -> Optional[Dict[str, object]]:
    cached = load_pitcher_stuff_comparison_cache_file()
    entries = cached.get("pitchers", {}) if isinstance(cached, dict) else {}
    if not isinstance(entries, dict):
        return None
    payload = entries.get(str(mlb_player_id))
    if isinstance(payload, dict):
        canonical_fangraphs_id = resolve_canonical_fangraphs_pitcher_id(mlb_player_id, payload.get("fangraphs_player_id"))
        if canonical_fangraphs_id and payload.get("fangraphs_player_id") != canonical_fangraphs_id:
            payload = {
                **payload,
                "fangraphs_player_id": canonical_fangraphs_id,
                "fangraphs_url": f"https://www.fangraphs.com/statss.aspx?playerid={canonical_fangraphs_id}",
            }
        return payload
    if fangraphs_player_id is not None:
        for entry in entries.values():
            if isinstance(entry, dict) and entry.get("fangraphs_player_id") == fangraphs_player_id:
                return entry
    return None


def get_or_fetch_pitcher_stuff_comparison(
    mlb_player_id: Optional[int],
    pitcher_name: str,
    fangraphs_player_id: Optional[int],
    current_year: int,
    previous_year: int,
) -> Optional[Dict[str, object]]:
    if not mlb_player_id:
        return None
    cached = get_cached_pitcher_stuff_comparison(mlb_player_id, fangraphs_player_id=fangraphs_player_id)
    if not pitcher_stuff_entry_needs_refresh(cached, current_year, previous_year):
        return cached
    if not pitcher_name or pitcher_name == "TBD":
        return cached
    try:
        incoming = fetch_pitcher_stuff_comparison_entry(
            mlb_player_id=mlb_player_id,
            pitcher_name=pitcher_name,
            fangraphs_player_id=fangraphs_player_id,
            current_year=current_year,
            previous_year=previous_year,
        )
    except Exception:
        return cached
    save_cached_pitcher_stuff_comparison(incoming)
    refreshed = get_cached_pitcher_stuff_comparison(mlb_player_id, fangraphs_player_id=fangraphs_player_id)
    return refreshed or incoming


def save_cached_pitcher_stuff_comparison(payload: Dict[str, object]) -> None:
    if not payload.get("mlb_player_id"):
        return
    canonical_fangraphs_id = resolve_canonical_fangraphs_pitcher_id(
        int(payload["mlb_player_id"]),
        payload.get("fangraphs_player_id"),
    )
    canonical_payload = {
        **payload,
        "fangraphs_player_id": canonical_fangraphs_id,
        "fangraphs_url": (
            f"https://www.fangraphs.com/statss.aspx?playerid={canonical_fangraphs_id}"
            if canonical_fangraphs_id
            else payload.get("fangraphs_url")
        ),
    }
    cached = load_pitcher_stuff_comparison_cache_file()
    entries = cached.setdefault("pitchers", {}) if isinstance(cached, dict) else {}
    if not isinstance(entries, dict):
        entries = {}
        cached["pitchers"] = entries
    existing_payload = entries.get(str(payload["mlb_player_id"]))
    entries[str(payload["mlb_player_id"])] = merge_pitcher_stuff_comparison_entries(existing_payload, canonical_payload)
    cached["generated_at"] = eastern_now().isoformat()
    save_pitcher_stuff_comparison_cache_file(cached)


def reconcile_pitcher_stuff_comparison_cache(payload: Dict[str, object]) -> Dict[str, object]:
    if not isinstance(payload, dict):
        return {"pitchers": {}}
    entries = payload.get("pitchers", {})
    if not isinstance(entries, dict):
        return {"pitchers": {}}

    reconciled_entries: Dict[str, Dict[str, object]] = {}
    for mlb_id, entry in entries.items():
        if not isinstance(entry, dict):
            continue
        try:
            mlb_int = int(entry.get("mlb_player_id") or mlb_id)
        except (TypeError, ValueError):
            continue
        canonical_fangraphs_id = resolve_canonical_fangraphs_pitcher_id(mlb_int, entry.get("fangraphs_player_id"))
        reconciled_entries[str(mlb_int)] = {
            **entry,
            "mlb_player_id": mlb_int,
            "fangraphs_player_id": canonical_fangraphs_id,
            "fangraphs_url": (
                f"https://www.fangraphs.com/statss.aspx?playerid={canonical_fangraphs_id}"
                if canonical_fangraphs_id
                else entry.get("fangraphs_url")
            ),
        }

    return {
        **payload,
        "generated_at": eastern_now().isoformat(),
        "pitchers": reconciled_entries,
    }


def extract_fangraphs_player_id_from_url(url: Optional[str]) -> Optional[int]:
    if not url:
        return None
    statss_match = re.search(r"[?&]playerid=(\d+)", url)
    if statss_match:
        return int(statss_match.group(1))
    path_match = re.search(r"/players/[^/]+/(\d+)/", url)
    if path_match:
        return int(path_match.group(1))
    return None


def build_velocity_pitcher_index(
    current_velocity_rows: List[Dict[str, object]],
    previous_velocity_rows: List[Dict[str, object]],
    probable_games: List[Dict[str, object]],
) -> Dict[str, Dict[str, object]]:
    pitcher_index: Dict[str, Dict[str, object]] = {}
    fangraphs_id_map = build_fangraphs_pitcher_id_map()

    for row in current_velocity_rows + previous_velocity_rows:
        pitcher_id = str(row.get("pitcher") or row.get("pitcher_id") or "").strip()
        if not pitcher_id:
            continue
        pitcher_entry = pitcher_index.setdefault(
            pitcher_id,
            {
                "pitcher_id": pitcher_id,
                "pitcher_name": str(row.get("name_display_first_last") or row.get("pitcher_name") or "").strip(),
                "fangraphs_player_id": resolve_canonical_fangraphs_pitcher_id(
                    int(pitcher_id) if pitcher_id.isdigit() else None,
                    row.get("fangraphs_player_id") or fangraphs_id_map.get(pitcher_id),
                ),
            },
        )
        if not pitcher_entry["pitcher_name"]:
            pitcher_entry["pitcher_name"] = str(row.get("name_display_first_last") or row.get("pitcher_name") or "").strip()
        if not pitcher_entry.get("fangraphs_player_id"):
            pitcher_entry["fangraphs_player_id"] = resolve_canonical_fangraphs_pitcher_id(
                int(pitcher_id) if pitcher_id.isdigit() else None,
                row.get("fangraphs_player_id") or fangraphs_id_map.get(pitcher_id),
            )

    for game in probable_games:
        for side in ("away_pitcher", "home_pitcher"):
            pitcher = game.get(side, {})
            pitcher_id = str(pitcher.get("mlb_id") or "").strip()
            if not pitcher_id:
                continue
            pitcher_entry = pitcher_index.setdefault(
                pitcher_id,
                {
                    "pitcher_id": pitcher_id,
                    "pitcher_name": pitcher.get("name", ""),
                    "fangraphs_player_id": None,
                },
            )
            if pitcher.get("name"):
                pitcher_entry["pitcher_name"] = pitcher["name"]
            if pitcher.get("fangraphs_player_id"):
                pitcher_entry["fangraphs_player_id"] = int(pitcher["fangraphs_player_id"])

    return pitcher_index


def build_league_wide_pitcher_index() -> Dict[str, Dict[str, object]]:
    pitcher_index: Dict[str, Dict[str, object]] = {}
    for team in TEAM_OPTIONS:
        try:
            roster_rows = fetch_depth_chart_data(team["slug"]).get("dataRoster", [])
        except Exception:
            roster_rows = []
        if not isinstance(roster_rows, list):
            continue
        for row in roster_rows:
            position_label = str(row.get("position1") or row.get("position") or "")
            roster_type = str(row.get("type") or "").lower()
            mlevel = str(row.get("mlevel") or "").upper()
            if "P" not in position_label:
                continue
            if not (roster_type.startswith("mlb-") or mlevel == "MLB"):
                continue
            mlb_id = row.get("mlbamid")
            fangraphs_id = row.get("playerid") or row.get("playerid1") or row.get("playerid2")
            if not mlb_id:
                continue
            try:
                mlb_int = int(mlb_id)
            except (TypeError, ValueError):
                continue
            pitcher_name = (
                row.get("playerName")
                or row.get("playername")
                or row.get("name")
                or row.get("fullName")
                or ""
            )
            pitcher_index[str(mlb_int)] = {
                "pitcher_id": str(mlb_int),
                "pitcher_name": str(pitcher_name).strip(),
                "fangraphs_player_id": resolve_canonical_fangraphs_pitcher_id(mlb_int, int(fangraphs_id) if fangraphs_id else None),
            }

    existing_pitch_stuff_cache = load_stuff_plus_cache_file()
    existing_pitchers = existing_pitch_stuff_cache.get("pitchers", {}) if isinstance(existing_pitch_stuff_cache, dict) else {}
    if isinstance(existing_pitchers, dict):
        for pitcher_id, entry in existing_pitchers.items():
            if not isinstance(entry, dict):
                continue
            try:
                mlb_int = int(entry.get("pitcher_id") or pitcher_id)
            except (TypeError, ValueError):
                continue
            existing_entry = pitcher_index.setdefault(
                str(mlb_int),
                {
                    "pitcher_id": str(mlb_int),
                    "pitcher_name": str(entry.get("pitcher_name") or "").strip(),
                    "fangraphs_player_id": None,
                },
            )
            if not existing_entry.get("pitcher_name") and entry.get("pitcher_name"):
                existing_entry["pitcher_name"] = str(entry.get("pitcher_name")).strip()
            canonical_fangraphs_id = resolve_canonical_fangraphs_pitcher_id(mlb_int, entry.get("fangraphs_player_id"))
            if canonical_fangraphs_id:
                existing_entry["fangraphs_player_id"] = canonical_fangraphs_id
    return pitcher_index


def build_pitch_stuff_plus_year_map(season_row: Optional[Dict[str, object]]) -> Dict[str, Optional[int]]:
    if not season_row:
        return {}
    return build_pitch_stuff_plus_map(season_row)


def count_nonempty_stuff_by_pitch(stuff_by_pitch: Dict[str, object]) -> int:
    return sum(1 for value in stuff_by_pitch.values() if value not in (None, ""))


def count_nonempty_pitch_stuff_values(years_payload: Dict[str, object]) -> int:
    count = 0
    for year_map in years_payload.values():
        if not isinstance(year_map, dict):
            continue
        count += sum(1 for value in year_map.values() if value not in (None, ""))
    return count


def build_fangraphs_player_page_candidates(
    pitcher_name: str,
    fangraphs_player_id: Optional[int] = None,
) -> List[Dict[str, object]]:
    candidates: List[Dict[str, object]] = []
    seen: set[str] = set()

    if fangraphs_player_id:
        url = f"https://www.fangraphs.com/statss.aspx?playerid={fangraphs_player_id}"
        if url not in seen:
            candidates.append({"url": url, "fangraphs_player_id": int(fangraphs_player_id), "source": "id"})
            seen.add(url)

    try:
        resolved_url = resolve_fangraphs_player_url(pitcher_name)
    except Exception:
        resolved_url = None
    resolved_id = extract_fangraphs_player_id_from_url(resolved_url) if resolved_url else None
    if resolved_url and resolved_url not in seen:
        candidates.append({"url": resolved_url, "fangraphs_player_id": resolved_id, "source": "name"})
        seen.add(resolved_url)

    return candidates


def merge_pitch_stuff_cache_entry(
    existing: Optional[Dict[str, object]],
    incoming: Dict[str, object],
) -> Dict[str, object]:
    if not existing:
        return incoming

    existing_years = existing.get("years", {}) if isinstance(existing.get("years"), dict) else {}
    incoming_years = incoming.get("years", {}) if isinstance(incoming.get("years"), dict) else {}
    merged_years: Dict[str, Dict[str, Optional[int]]] = {}
    for year_key in sorted(set(existing_years.keys()) | set(incoming_years.keys())):
        existing_map = existing_years.get(year_key, {}) if isinstance(existing_years.get(year_key), dict) else {}
        incoming_map = incoming_years.get(year_key, {}) if isinstance(incoming_years.get(year_key), dict) else {}
        merged_years[year_key] = {
            pitch_key: incoming_map.get(pitch_key) if incoming_map.get(pitch_key) not in (None, "") else existing_map.get(pitch_key)
            for pitch_key in sorted(set(existing_map.keys()) | set(incoming_map.keys()))
        }

    merged = {
        **existing,
        **incoming,
        "pitcher_name": incoming.get("pitcher_name") or existing.get("pitcher_name"),
        "fangraphs_player_id": incoming.get("fangraphs_player_id") or existing.get("fangraphs_player_id"),
        "years": merged_years,
    }
    return merged


def repair_saved_fangraphs_ids_in_caches() -> Dict[str, object]:
    repaired_pitch_cache = load_stuff_plus_cache_file()
    repaired_pitchers = repaired_pitch_cache.get("pitchers", {}) if isinstance(repaired_pitch_cache, dict) else {}
    repaired_count = 0

    if isinstance(repaired_pitchers, dict):
        for pitcher_id, entry in repaired_pitchers.items():
            if not isinstance(entry, dict):
                continue
            try:
                mlb_int = int(entry.get("pitcher_id") or pitcher_id)
            except (TypeError, ValueError):
                continue
            pitcher_name = str(entry.get("pitcher_name") or "").strip()
            canonical_id = resolve_canonical_fangraphs_pitcher_id(mlb_int, entry.get("fangraphs_player_id"))
            years_payload = entry.get("years", {}) if isinstance(entry.get("years"), dict) else {}
            should_name_repair = (
                not canonical_id
                or count_nonempty_pitch_stuff_values(years_payload) == 0
            )
            try:
                resolved_url = resolve_fangraphs_player_url(pitcher_name) if pitcher_name and should_name_repair else None
            except Exception:
                resolved_url = None
            resolved_id = extract_fangraphs_player_id_from_url(resolved_url) if resolved_url else None
            best_id = resolved_id or canonical_id or entry.get("fangraphs_player_id")
            if best_id != entry.get("fangraphs_player_id"):
                repaired_count += 1
            entry["fangraphs_player_id"] = best_id
            if pitcher_name:
                entry["pitcher_name"] = pitcher_name

    repaired_pitch_cache["generated_at"] = eastern_now().isoformat()
    save_stuff_plus_cache_file(repaired_pitch_cache)

    repaired_profile_cache = reconcile_pitcher_stuff_comparison_cache(load_pitcher_stuff_comparison_cache_file())
    profile_pitchers = repaired_profile_cache.get("pitchers", {}) if isinstance(repaired_profile_cache, dict) else {}
    profile_repaired_count = 0
    if isinstance(profile_pitchers, dict):
        for pitcher_id, entry in profile_pitchers.items():
            if not isinstance(entry, dict):
                continue
            pitch_entry = repaired_pitchers.get(str(pitcher_id)) if isinstance(repaired_pitchers, dict) else None
            pitch_fangraphs_id = pitch_entry.get("fangraphs_player_id") if isinstance(pitch_entry, dict) else None
            try:
                mlb_int = int(entry.get("mlb_player_id") or pitcher_id)
            except (TypeError, ValueError):
                continue
            canonical_id = resolve_canonical_fangraphs_pitcher_id(mlb_int, pitch_fangraphs_id or entry.get("fangraphs_player_id"))
            if canonical_id != entry.get("fangraphs_player_id"):
                profile_repaired_count += 1
            entry["fangraphs_player_id"] = canonical_id
            if canonical_id:
                entry["fangraphs_url"] = f"https://www.fangraphs.com/statss.aspx?playerid={canonical_id}"
    repaired_profile_cache["generated_at"] = eastern_now().isoformat()
    save_pitcher_stuff_comparison_cache_file(repaired_profile_cache)

    return {
        "pitch_cache_repairs": repaired_count,
        "profile_cache_repairs": profile_repaired_count,
        "pitch_cache_pitchers": len(repaired_pitchers) if isinstance(repaired_pitchers, dict) else 0,
        "profile_cache_pitchers": len(profile_pitchers) if isinstance(profile_pitchers, dict) else 0,
    }


def fetch_pitcher_stuff_comparison_entry(
    mlb_player_id: int,
    pitcher_name: str,
    fangraphs_player_id: Optional[int] = None,
    current_year: Optional[int] = None,
    previous_year: Optional[int] = None,
) -> Dict[str, object]:
    current_year = current_year or eastern_now().year
    previous_year = previous_year or (current_year - 1)
    canonical_fangraphs_player_id = resolve_canonical_fangraphs_pitcher_id(mlb_player_id, fangraphs_player_id)
    best_payload: Optional[Dict[str, object]] = None
    best_score = -1
    last_error: Optional[Exception] = None

    for candidate in build_fangraphs_player_page_candidates(pitcher_name, canonical_fangraphs_player_id):
        try:
            fg_html = fetch_fangraphs_page(str(candidate["url"]))
            fg_payload = extract_next_data(fg_html)
            page_props = fg_payload.get("props", {}).get("pageProps", {})
            season_rows = collect_pitching_rows(page_props)
            current_season_row = find_pitching_row_for_season(season_rows, current_year)
            previous_season_row = find_pitching_row_for_season(season_rows, previous_year)
            current_stuff_value = current_season_row.get("sp_stuff") if current_season_row else None
            previous_stuff_value = previous_season_row.get("sp_stuff") if previous_season_row else None
            try:
                stuff_diff_value = (
                    float(current_stuff_value) - float(previous_stuff_value)
                    if current_stuff_value not in (None, "") and previous_stuff_value not in (None, "")
                    else None
                )
            except (TypeError, ValueError):
                stuff_diff_value = None
            location_plus_value = current_season_row.get("sp_location") if current_season_row else None
            pitching_plus_value = current_season_row.get("sp_pitching") if current_season_row else None
            candidate_payload = build_pitcher_stuff_comparison_payload(
                mlb_player_id=mlb_player_id,
                fangraphs_player_id=candidate.get("fangraphs_player_id"),
                pitcher_name=pitcher_name,
                fangraphs_url=str(candidate["url"]),
                current_year=current_year,
                previous_year=previous_year,
                current_stuff_value=current_stuff_value,
                previous_stuff_value=previous_stuff_value,
                stuff_diff_value=stuff_diff_value,
                location_plus_value=location_plus_value,
                pitching_plus_value=pitching_plus_value,
            )
            candidate_metrics = candidate_payload.get("metrics", {})
            candidate_score = count_nonempty_pitcher_stuff_metrics(candidate_metrics if isinstance(candidate_metrics, dict) else {})
            if candidate_score > best_score:
                best_payload = candidate_payload
                best_score = candidate_score
            if candidate_score >= 3:
                break
        except Exception as exc:
            last_error = exc

    if best_payload:
        return best_payload
    if last_error:
        raise last_error
    raise ValueError("Could not resolve a Fangraphs player page for this pitcher.")


def fetch_pitch_stuff_plus_years(
    pitcher_name: str,
    seasons: List[int],
    fangraphs_player_id: Optional[int] = None,
) -> Dict[str, object]:
    best_payload: Optional[Dict[str, object]] = None
    best_score = -1
    last_error: Optional[Exception] = None

    for candidate in build_fangraphs_player_page_candidates(pitcher_name, fangraphs_player_id):
        try:
            fg_html = fetch_fangraphs_page(str(candidate["url"]))
            fg_payload = extract_next_data(fg_html)
            data_stats = fg_payload.get("props", {}).get("pageProps", {}).get("dataStats", {})
            season_rows = data_stats.get("data", [])
            years_payload = {
                str(season): build_pitch_stuff_plus_year_map(find_pitching_row_for_season(season_rows, season))
                for season in seasons
            }
            candidate_payload = {
                "pitcher_name": pitcher_name,
                "fangraphs_player_id": candidate.get("fangraphs_player_id"),
                "years": years_payload,
            }
            candidate_score = count_nonempty_pitch_stuff_values(years_payload)
            if candidate_score > best_score:
                best_payload = candidate_payload
                best_score = candidate_score
            if candidate_score > 0:
                break
        except Exception as exc:
            last_error = exc

    if best_payload:
        return best_payload
    if last_error:
        raise last_error
    return {
        "pitcher_name": pitcher_name,
        "fangraphs_player_id": fangraphs_player_id,
        "years": {str(season): {} for season in seasons},
    }


def rebuild_pitcher_stuff_comparison_cache_league_wide() -> Dict[str, object]:
    pitcher_index = build_league_wide_pitcher_index()
    existing_cache = reconcile_pitcher_stuff_comparison_cache(load_pitcher_stuff_comparison_cache_file())
    existing_entries = existing_cache.get("pitchers", {}) if isinstance(existing_cache, dict) else {}
    rebuilt_entries: Dict[str, Dict[str, object]] = {}
    current_year = eastern_now().year
    previous_year = current_year - 1
    progress_since_save = 0

    for pitcher_id, pitcher_info in pitcher_index.items():
        existing_entry = existing_entries.get(pitcher_id)
        if pitcher_stuff_entry_has_useful_values(existing_entry, current_year, previous_year):
            rebuilt_entries[pitcher_id] = existing_entry
            continue
        try:
            incoming_entry = fetch_pitcher_stuff_comparison_entry(
                mlb_player_id=int(pitcher_id),
                pitcher_name=pitcher_info.get("pitcher_name", ""),
                fangraphs_player_id=pitcher_info.get("fangraphs_player_id"),
                current_year=current_year,
                previous_year=previous_year,
            )
            rebuilt_entries[pitcher_id] = merge_pitcher_stuff_comparison_entries(existing_entry, incoming_entry)
        except Exception:
            if existing_entry:
                rebuilt_entries[pitcher_id] = existing_entry
            else:
                rebuilt_entries[pitcher_id] = build_pitcher_stuff_comparison_payload(
                    mlb_player_id=int(pitcher_id),
                    fangraphs_player_id=pitcher_info.get("fangraphs_player_id"),
                    pitcher_name=pitcher_info.get("pitcher_name", ""),
                    fangraphs_url=(
                        f"https://www.fangraphs.com/statss.aspx?playerid={pitcher_info.get('fangraphs_player_id')}"
                        if pitcher_info.get("fangraphs_player_id")
                        else None
                    ),
                    current_year=current_year,
                    previous_year=previous_year,
                    current_stuff_value=None,
                    previous_stuff_value=None,
                    stuff_diff_value=None,
                    location_plus_value=None,
                    pitching_plus_value=None,
                )
        progress_since_save += 1
        if progress_since_save >= 25:
            interim_payload = reconcile_pitcher_stuff_comparison_cache(
                {
                    "generated_at": eastern_now().isoformat(),
                    "pitchers": {
                        **existing_entries,
                        **rebuilt_entries,
                    },
                }
            )
            save_pitcher_stuff_comparison_cache_file(interim_payload)
            progress_since_save = 0
        time.sleep(0.2)

    for pitcher_id, existing_entry in existing_entries.items():
        if pitcher_id not in rebuilt_entries:
            rebuilt_entries[pitcher_id] = existing_entry

    rebuilt_payload = reconcile_pitcher_stuff_comparison_cache(
        {
            "generated_at": eastern_now().isoformat(),
            "pitchers": rebuilt_entries,
        }
    )
    save_pitcher_stuff_comparison_cache_file(rebuilt_payload)
    return rebuilt_payload


def build_daily_stuff_plus_cache(
    pitcher_index: Dict[str, Dict[str, object]],
    seasons: List[int],
) -> Dict[str, object]:
    existing_cache = load_stuff_plus_cache_file()
    existing_pitchers = existing_cache.get("pitchers", {}) if isinstance(existing_cache, dict) else {}
    pitchers_payload: Dict[str, Dict[str, object]] = {}

    for pitcher_id, pitcher_info in pitcher_index.items():
        existing_entry = existing_pitchers.get(pitcher_id) if isinstance(existing_pitchers, dict) else None
        if isinstance(existing_entry, dict):
            existing_years = existing_entry.get("years", {})
            if (
                existing_entry.get("fangraphs_player_id")
                and count_nonempty_pitch_stuff_values(existing_years if isinstance(existing_years, dict) else {}) >= 6
            ):
                pitchers_payload[pitcher_id] = existing_entry
                continue

        payload: Dict[str, object]
        try:
            payload = fetch_pitch_stuff_plus_years(
                pitcher_info.get("pitcher_name", ""),
                seasons=seasons,
                fangraphs_player_id=pitcher_info.get("fangraphs_player_id"),
            )
        except Exception:
            time.sleep(0.75)
            try:
                payload = fetch_pitch_stuff_plus_years(
                    pitcher_info.get("pitcher_name", ""),
                    seasons=seasons,
                    fangraphs_player_id=pitcher_info.get("fangraphs_player_id"),
                )
            except Exception:
                payload = {
                    "pitcher_name": pitcher_info.get("pitcher_name", ""),
                    "fangraphs_player_id": pitcher_info.get("fangraphs_player_id"),
                    "years": {str(season): {} for season in seasons},
                }
        incoming_entry = {
            "pitcher_id": pitcher_id,
            "pitcher_name": payload.get("pitcher_name", pitcher_info.get("pitcher_name", "")),
            "fangraphs_player_id": payload.get("fangraphs_player_id"),
            "years": payload.get("years", {}),
        }
        pitchers_payload[pitcher_id] = merge_pitch_stuff_cache_entry(existing_entry if isinstance(existing_entry, dict) else None, incoming_entry)
        time.sleep(0.2)

    if isinstance(existing_pitchers, dict):
        for pitcher_id, existing_entry in existing_pitchers.items():
            if pitcher_id not in pitchers_payload and isinstance(existing_entry, dict):
                pitchers_payload[pitcher_id] = existing_entry

    return {
        "generated_at": eastern_now().isoformat(),
        "refresh_key": current_stuff_plus_refresh_key(),
        "seasons": seasons,
        "pitchers": pitchers_payload,
    }


def start_background_stuff_plus_refresh(
    pitcher_index: Dict[str, Dict[str, object]],
    seasons: List[int],
) -> None:
    global STUFF_PLUS_REFRESH_IN_PROGRESS

    with STUFF_PLUS_REFRESH_LOCK:
        if STUFF_PLUS_REFRESH_IN_PROGRESS:
            return
        STUFF_PLUS_REFRESH_IN_PROGRESS = True

    def runner() -> None:
        global STUFF_PLUS_REFRESH_IN_PROGRESS
        try:
            refreshed = build_daily_stuff_plus_cache(pitcher_index, seasons)
            save_stuff_plus_cache_file(refreshed)
        except Exception:
            pass
        finally:
            with STUFF_PLUS_REFRESH_LOCK:
                STUFF_PLUS_REFRESH_IN_PROGRESS = False

    threading.Thread(target=runner, name="stuff-plus-refresh", daemon=True).start()


def get_or_build_daily_stuff_plus_cache(
    pitcher_index: Dict[str, Dict[str, object]],
    seasons: List[int],
) -> Dict[str, object]:
    expected_refresh_key = current_stuff_plus_refresh_key()
    cached = load_stuff_plus_cache_file()
    cached_pitchers = cached.get("pitchers", {}) if isinstance(cached, dict) else {}
    cached_seasons = cached.get("seasons", []) if isinstance(cached, dict) else []

    cache_nonempty_count = sum(
        any(year_map for year_map in (pitcher_payload.get("years", {}) or {}).values())
        for pitcher_payload in cached_pitchers.values()
    )
    cache_has_values = cache_nonempty_count >= max(1, len(pitcher_index) // 2)

    if (
        cached.get("refresh_key") == expected_refresh_key
        and cached_seasons == seasons
        and cache_has_values
        and all(pitcher_id in cached_pitchers for pitcher_id in pitcher_index)
    ):
        return cached

    if cached_pitchers:
        start_background_stuff_plus_refresh(pitcher_index, seasons)
        return cached

    refreshed = build_daily_stuff_plus_cache(pitcher_index, seasons)
    save_stuff_plus_cache_file(refreshed)
    return refreshed


def attach_stuff_plus_from_cache(
    rows: List[Dict[str, object]],
    stuff_cache: Dict[str, object],
    current_year: int,
    previous_year: int,
) -> None:
    pitchers_payload = stuff_cache.get("pitchers", {}) if isinstance(stuff_cache, dict) else {}
    pitchers_by_fangraphs_id: Dict[int, Dict[str, object]] = {}
    pitchers_by_name: Dict[str, Dict[str, object]] = {}

    for pitcher_payload in pitchers_payload.values():
        if not isinstance(pitcher_payload, dict):
            continue
        fangraphs_id = pitcher_payload.get("fangraphs_player_id")
        if fangraphs_id:
            try:
                pitchers_by_fangraphs_id[int(fangraphs_id)] = pitcher_payload
            except (TypeError, ValueError):
                pass
        normalized_pitcher_name = normalize_name(str(pitcher_payload.get("pitcher_name") or ""))
        if normalized_pitcher_name:
            pitchers_by_name[normalized_pitcher_name] = pitcher_payload

    for row in rows:
        pitcher_id = str(row.get("pitcher_id") or "")
        pitcher_payload = pitchers_payload.get(pitcher_id)
        if not pitcher_payload:
            canonical_fangraphs_id = resolve_canonical_fangraphs_pitcher_id(
                int(pitcher_id) if pitcher_id.isdigit() else None,
                row.get("fangraphs_player_id"),
            )
            if canonical_fangraphs_id:
                pitcher_payload = pitchers_by_fangraphs_id.get(canonical_fangraphs_id)
        if not pitcher_payload:
            pitcher_payload = pitchers_by_name.get(normalize_name(str(row.get("pitcher_name") or "")))
        years_payload = pitcher_payload.get("years", {}) if pitcher_payload else {}
        current_map = years_payload.get(str(current_year), {})
        previous_map = years_payload.get(str(previous_year), {})
        pitch_code = str(row.get("pitch_code", "")).upper()
        row["stuff_plus"] = current_map.get(pitch_code)
        row["previous_stuff_plus"] = previous_map.get(pitch_code)


def parse_float(value: object) -> Optional[float]:
    if value in (None, "", "null"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def fetch_savant_payloads(savant_sources: Dict[str, str]) -> Dict[str, List[Dict[str, object]]]:
    try:
        with ThreadPoolExecutor(max_workers=4) as executor:
            future_map = {
                source_key: executor.submit(fetch_text_page, source_url)
                for source_key, source_url in savant_sources.items()
            }
            return {
                source_key: extract_savant_pitch_data(future.result())
                for source_key, future in future_map.items()
            }
    except Exception:
        # Baseball Savant can reject bursty parallel requests, so fall back to sequential fetches.
        return {
            source_key: extract_savant_pitch_data(fetch_text_page(source_url))
            for source_key, source_url in savant_sources.items()
        }


def build_velocity_comparison() -> Dict[str, object]:
    current_year = eastern_now().year
    previous_year = current_year - 1
    current_velocity_url = SAVANT_PITCH_ARSENALS_URL.format(year=current_year, minimum="q", metric_type="avg_speed")
    previous_velocity_url = SAVANT_PITCH_ARSENALS_URL.format(year=previous_year, minimum="100", metric_type="avg_speed")
    current_spin_url = SAVANT_PITCH_ARSENALS_URL.format(year=current_year, minimum="q", metric_type="avg_spin")
    previous_spin_url = SAVANT_PITCH_ARSENALS_URL.format(year=previous_year, minimum="100", metric_type="avg_spin")
    current_usage_url = SAVANT_PITCH_ARSENALS_URL.format(year=current_year, minimum="q", metric_type="n_")
    previous_usage_url = SAVANT_PITCH_ARSENALS_URL.format(year=previous_year, minimum="100", metric_type="n_")

    savant_sources = {
        "current_velocity": current_velocity_url,
        "previous_velocity": previous_velocity_url,
        "current_spin": current_spin_url,
        "previous_spin": previous_spin_url,
        "current_usage": current_usage_url,
        "previous_usage": previous_usage_url,
    }

    savant_payloads = fetch_savant_payloads(savant_sources)

    current_velocity_rows = savant_payloads["current_velocity"]
    previous_velocity_rows = savant_payloads["previous_velocity"]
    current_spin_rows = savant_payloads["current_spin"]
    previous_spin_rows = savant_payloads["previous_spin"]
    current_usage_rows = savant_payloads["current_usage"]
    previous_usage_rows = savant_payloads["previous_usage"]

    previous_velocity_by_pitcher = {str(row.get("pitcher")): row for row in previous_velocity_rows}
    current_spin_by_pitcher = {str(row.get("pitcher")): row for row in current_spin_rows}
    previous_spin_by_pitcher = {str(row.get("pitcher")): row for row in previous_spin_rows}
    current_usage_by_pitcher = {str(row.get("pitcher")): row for row in current_usage_rows}
    previous_usage_by_pitcher = {str(row.get("pitcher")): row for row in previous_usage_rows}
    comparison_rows: List[Dict[str, object]] = []
    team_labels = set()

    for current_velocity in current_velocity_rows:
        pitcher_id = str(current_velocity.get("pitcher") or "").strip()
        if not pitcher_id:
            continue
        previous_velocity = previous_velocity_by_pitcher.get(pitcher_id, {})
        current_spin = current_spin_by_pitcher.get(pitcher_id, {})
        previous_spin = previous_spin_by_pitcher.get(pitcher_id, {})
        current_usage = current_usage_by_pitcher.get(pitcher_id, {})
        previous_usage = previous_usage_by_pitcher.get(pitcher_id, {})
        team = str(current_velocity.get("name_abbrev") or current_velocity.get("name_display_club") or "").strip()
        if team:
            team_labels.add(team)

        for pitch_code, pitch_name in SAVANT_PITCH_TYPES:
            current_velocity_value = parse_float(current_velocity.get(f"{pitch_code}_avg_speed"))
            if current_velocity_value is None:
                continue

            previous_velocity_value = parse_float(previous_velocity.get(f"{pitch_code}_avg_speed"))
            velocity_delta = round(current_velocity_value - previous_velocity_value, 1) if previous_velocity_value is not None else None
            current_spin_value = parse_float(current_spin.get(f"{pitch_code}_avg_spin"))
            previous_spin_value = parse_float(previous_spin.get(f"{pitch_code}_avg_spin"))
            spin_delta = round(current_spin_value - previous_spin_value, 0) if previous_spin_value is not None and current_spin_value is not None else None
            usage_value = parse_float(current_usage.get(f"n_{pitch_code}"))
            previous_usage_value = parse_float(previous_usage.get(f"n_{pitch_code}"))
            comparison_rows.append(
                {
                    "pitcher_id": pitcher_id,
                    "pitcher_name": str(current_velocity.get("name_display_first_last") or "").strip(),
                    "team": team,
                    "throws": str(current_velocity.get("pitch_hand") or "").strip(),
                    "pitch_type": pitch_name,
                    "pitch_code": pitch_code.upper(),
                    "current_velocity": round(current_velocity_value, 1),
                    "previous_velocity": round(previous_velocity_value, 1) if previous_velocity_value is not None else None,
                    "delta": velocity_delta,
                    "current_spin": round(current_spin_value, 0) if current_spin_value is not None else None,
                    "previous_spin": round(previous_spin_value, 0) if previous_spin_value is not None else None,
                    "spin_delta": int(spin_delta) if spin_delta is not None else None,
                    "usage": round(usage_value, 1) if usage_value is not None else None,
                    "previous_usage": round(previous_usage_value, 1) if previous_usage_value is not None else None,
                    "usage_delta": round(usage_value - previous_usage_value, 1) if usage_value is not None and previous_usage_value is not None else None,
                }
            )

    comparison_rows.sort(
        key=lambda row: (
            -(abs(row["delta"]) if row["delta"] is not None else -1),
            row["pitcher_name"],
            row["pitch_type"],
        )
    )

    rows_by_pitcher: Dict[str, List[Dict[str, object]]] = {}
    for row in comparison_rows:
        rows_by_pitcher.setdefault(row["pitcher_id"], []).append(row)

    probable_date = next_mlb_date()
    probable_payload = fetch_probable_pitchers(probable_date)
    pitcher_index = build_velocity_pitcher_index(current_velocity_rows, previous_velocity_rows, probable_payload.get("games", []))
    stuff_plus_cache = get_or_build_daily_stuff_plus_cache(pitcher_index, [previous_year, current_year])
    attach_stuff_plus_from_cache(comparison_rows, stuff_plus_cache, current_year, previous_year)
    tomorrow_games: List[Dict[str, object]] = []

    for game in probable_payload.get("games", []):
        starters = []
        for side, team_key in (("away_pitcher", "away_team"), ("home_pitcher", "home_team")):
            pitcher = game.get(side, {})
            mlb_id = pitcher.get("mlb_id")
            pitcher_rows = rows_by_pitcher.get(str(mlb_id), []) if mlb_id else []
            pitcher_stuff_entry = get_or_fetch_pitcher_stuff_comparison(
                mlb_id,
                pitcher.get("name", "TBD"),
                pitcher.get("fangraphs_player_id"),
                current_year,
                previous_year,
            )
            pitcher_stuff_metrics = pitcher_stuff_entry.get("metrics", {}) if isinstance(pitcher_stuff_entry, dict) else {}
            starters.append(
                {
                    "team": game.get(team_key, ""),
                    "pitcher_name": pitcher.get("name", "TBD"),
                    "hand": pitcher.get("hand", ""),
                    "record": pitcher.get("record", ""),
                    "era": pitcher.get("era", ""),
                    "mlb_id": mlb_id,
                    "fangraphs_player_id": pitcher.get("fangraphs_player_id"),
                    "rows": pitcher_rows,
                    "pitcher_stuff_summary": {
                        "current": pitcher_stuff_metrics.get("StuffPlusCurrent"),
                        "previous": pitcher_stuff_metrics.get("StuffPlusPrevious"),
                        "diff": pitcher_stuff_metrics.get("StuffPlusDiff"),
                        "saved_at": (pitcher_stuff_entry or {}).get("cache_saved_at") if isinstance(pitcher_stuff_entry, dict) else None,
                    },
                }
            )

        tomorrow_games.append(
            {
                "game_time": game.get("game_time", ""),
                "away_team": game.get("away_team", ""),
                "home_team": game.get("home_team", ""),
                "starters": starters,
            }
        )

    return {
        "current_year": current_year,
        "previous_year": previous_year,
        "source_urls": {
            "current_velocity": current_velocity_url,
            "previous_velocity": previous_velocity_url,
            "current_spin": current_spin_url,
            "previous_spin": previous_spin_url,
        },
        "stuff_plus_cache": {
            "path": str(STUFF_PLUS_CACHE_PATH),
            "generated_at": stuff_plus_cache.get("generated_at"),
            "refresh_key": stuff_plus_cache.get("refresh_key"),
        },
        "tomorrow": {
            "date": probable_payload.get("date", probable_date.isoformat()),
            "source_url": probable_payload.get("source_url", f"{MLB_SCORES_URL}/{probable_date.isoformat()}"),
            "games": tomorrow_games,
        },
        "refreshed_every_seconds": REFRESH_SECONDS,
        "teams": sorted(team_labels),
        "rows": comparison_rows,
    }


def get_or_build_velocity_comparison_cache() -> Dict[str, object]:
    expected_refresh_key = current_stuff_plus_refresh_key()
    current_timestamp = time.time()

    with VELOCITY_RUNTIME_CACHE_LOCK:
        if (
            VELOCITY_RUNTIME_CACHE.get("refresh_key") == expected_refresh_key
            and VELOCITY_RUNTIME_CACHE.get("expires_at", 0) > current_timestamp
            and VELOCITY_RUNTIME_CACHE.get("payload")
        ):
            runtime_payload = dict(VELOCITY_RUNTIME_CACHE["payload"])
            runtime_payload["cache_generated_at"] = VELOCITY_RUNTIME_CACHE.get("generated_at")
            return runtime_payload

    cached = load_velocity_comparison_cache_file()

    if cached.get("refresh_key") == expected_refresh_key and cached.get("payload"):
        with VELOCITY_RUNTIME_CACHE_LOCK:
            VELOCITY_RUNTIME_CACHE.clear()
            VELOCITY_RUNTIME_CACHE.update(
                {
                    "refresh_key": expected_refresh_key,
                    "generated_at": cached.get("generated_at"),
                    "expires_at": current_timestamp + VELOCITY_RUNTIME_CACHE_SECONDS,
                    "payload": cached["payload"],
                }
            )
        cached_payload = dict(cached["payload"])
        cached_payload["cache_generated_at"] = cached.get("generated_at")
        return cached_payload

    try:
        payload = build_velocity_comparison()
    except Exception:
        if cached.get("payload"):
            with VELOCITY_RUNTIME_CACHE_LOCK:
                VELOCITY_RUNTIME_CACHE.clear()
                VELOCITY_RUNTIME_CACHE.update(
                    {
                        "refresh_key": cached.get("refresh_key", expected_refresh_key),
                        "generated_at": cached.get("generated_at"),
                        "expires_at": current_timestamp + VELOCITY_RUNTIME_CACHE_SECONDS,
                        "payload": cached["payload"],
                    }
                )
            cached_payload = dict(cached["payload"])
            cached_payload["cache_generated_at"] = cached.get("generated_at")
            return cached_payload
        raise

    generated_at = eastern_now().isoformat()
    cache_wrapper = {
        "generated_at": generated_at,
        "refresh_key": expected_refresh_key,
        "payload": payload,
    }
    save_velocity_comparison_cache_file(cache_wrapper)
    with VELOCITY_RUNTIME_CACHE_LOCK:
        VELOCITY_RUNTIME_CACHE.clear()
        VELOCITY_RUNTIME_CACHE.update(
            {
                "refresh_key": expected_refresh_key,
                "generated_at": generated_at,
                "expires_at": current_timestamp + VELOCITY_RUNTIME_CACHE_SECONDS,
                "payload": payload,
            }
        )
    payload_with_cache = dict(payload)
    payload_with_cache["cache_generated_at"] = generated_at
    return payload_with_cache


def build_pitch_stuff_plus_map(current_row: Dict[str, object]) -> Dict[str, Optional[int]]:
    pitch_map: Dict[str, Optional[int]] = {}
    for savant_code, fg_code in STUFF_PLUS_FIELD_MAP.items():
        if not fg_code:
            pitch_map[savant_code] = None
            continue
        value = parse_float(current_row.get(f"sp_s_{fg_code}"))
        pitch_map[savant_code] = int(round(value)) if value is not None else None
    return pitch_map


def fetch_pitch_stuff_plus(
    pitcher_name: str,
    mlb_id: Optional[int] = None,
    fangraphs_player_id: Optional[int] = None,
) -> Dict[str, object]:
    canonical_fangraphs_player_id = resolve_canonical_fangraphs_pitcher_id(mlb_id, fangraphs_player_id)
    cache_key = f"{canonical_fangraphs_player_id or ''}|{mlb_id or ''}|{normalize_name(pitcher_name)}"
    if cache_key in PITCH_STUFF_CACHE:
        return PITCH_STUFF_CACHE[cache_key]

    best_payload: Optional[Dict[str, object]] = None
    best_score = -1
    last_error: Optional[Exception] = None

    for candidate in build_fangraphs_player_page_candidates(pitcher_name, canonical_fangraphs_player_id):
        try:
            fg_html = fetch_fangraphs_page(str(candidate["url"]))
            fg_payload = extract_next_data(fg_html)
            data_common = fg_payload.get("props", {}).get("pageProps", {}).get("dataCommon", {})
            current_row = find_current_pitching_row(data_common)
            candidate_payload = {
                "pitcher_name": pitcher_name,
                "stuff_by_pitch": build_pitch_stuff_plus_map(current_row) if current_row else {},
                "fangraphs_player_id": candidate.get("fangraphs_player_id"),
            }
            candidate_score = count_nonempty_stuff_by_pitch(candidate_payload["stuff_by_pitch"])
            if candidate_score > best_score:
                best_payload = candidate_payload
                best_score = candidate_score
            if candidate_score > 0:
                break
        except Exception as exc:
            last_error = exc

    payload = best_payload or {"pitcher_name": pitcher_name, "stuff_by_pitch": {}, "fangraphs_player_id": canonical_fangraphs_player_id}
    if not best_payload and last_error:
        payload["error"] = str(last_error)
    PITCH_STUFF_CACHE[cache_key] = payload
    return payload


def attach_pitch_stuff_plus(rows: List[Dict[str, object]], pitcher_name: str, mlb_id: Optional[int] = None, fangraphs_player_id: Optional[int] = None) -> None:
    if not rows:
        return
    stuff_payload = fetch_pitch_stuff_plus(pitcher_name, mlb_id=mlb_id, fangraphs_player_id=fangraphs_player_id)
    stuff_by_pitch = stuff_payload.get("stuff_by_pitch", {})
    for row in rows:
        row["stuff_plus"] = stuff_by_pitch.get(str(row.get("pitch_code", "")).upper())


def build_bullpen_excel_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bullpens"

    raw_html = fetch_fangraphs_page(BULLPEN_URL)
    payload = parse_bullpen_page(raw_html)
    payload = save_cached_bullpen_data(payload)

    for index, team in enumerate(TEAM_OPTIONS):
        start_col = 1 + index * 4
        team_payload = payload["teams"][team["slug"]]
        rows = team_payload["active_rows"]

        ws.cell(row=1, column=start_col, value=team["code"])
        ws.cell(row=2, column=start_col, value="Name")
        ws.cell(row=2, column=start_col + 1, value="Throws")
        ws.cell(row=2, column=start_col + 2, value="Role")

        for row_index, row in enumerate(rows, start=3):
            ws.cell(row=row_index, column=start_col, value=row["name"])
            ws.cell(row=row_index, column=start_col + 1, value=row["throws"])
            ws.cell(row=row_index, column=start_col + 2, value=row["role"])

        ws.column_dimensions[get_column_letter(start_col)].width = 24
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 10
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 22
        ws.column_dimensions[get_column_letter(start_col + 3)].width = 4

    ws.freeze_panes = "A1"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def team_block_height(payload: Dict[str, object]) -> int:
    layout = payload["excel_layout"]
    vsr = layout["vsR"]
    vsl = layout["vsL"]
    return max(
        1 + len(vsr["starters"]) + 2 + 1 + len(vsr["bench"]) + 2 + 1 + len(vsr["il"]),
        1 + len(vsl["starters"]),
    )


def write_team_block(ws, start_row: int, start_col: int, payload: Dict[str, object]) -> None:
    code = payload["team"]["code"]
    layout = payload["excel_layout"]
    vsr = layout["vsR"]
    vsl = layout["vsL"]

    ws.cell(row=start_row, column=start_col, value=code)
    ws.cell(row=start_row, column=start_col + 1, value=f"{code}L")

    current_row = start_row + 1
    for idx in range(max(len(vsr["starters"]), len(vsl["starters"]))):
        if idx < len(vsr["starters"]):
            ws.cell(row=current_row + idx, column=start_col, value=vsr["starters"][idx])
        if idx < len(vsl["starters"]):
            ws.cell(row=current_row + idx, column=start_col + 1, value=vsl["starters"][idx])

    current_row += max(len(vsr["starters"]), len(vsl["starters"]))
    current_row += 2

    ws.cell(row=current_row, column=start_col, value="Bench")
    current_row += 1
    for name in vsr["bench"]:
        ws.cell(row=current_row, column=start_col, value=name)
        current_row += 1

    current_row += 2
    ws.cell(row=current_row, column=start_col, value="IL")
    current_row += 1
    for name in vsr["il"]:
        ws.cell(row=current_row, column=start_col, value=name)
        current_row += 1


def build_excel_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Platoon Lineups"

    for index, team in enumerate(TEAM_OPTIONS):
        start_col = 1 + index * 3
        payload = fetch_lineup_data(team["slug"])
        write_team_block(ws, 1, start_col, payload)
        ws.column_dimensions[get_column_letter(start_col)].width = 24
        ws.column_dimensions[get_column_letter(start_col + 1)].width = 24
        ws.column_dimensions[get_column_letter(start_col + 2)].width = 4

    ws.freeze_panes = "A1"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


app = FastAPI(title="MLB Pricing Tools", version="1.4.0")


@app.on_event("startup")
def startup_homepage_refresh() -> None:
    ensure_homepage_cache_refresh_thread()


@app.get("/", response_class=HTMLResponse)
def home() -> HTMLResponse:
    return HTMLResponse(HOME_TEMPLATE_PATH.read_text(encoding="utf-8"))


@app.get("/platoon-lineups", response_class=HTMLResponse)
def platoon_lineups_page() -> HTMLResponse:
    return HTMLResponse(LINEUP_TEMPLATE_PATH.read_text(encoding="utf-8"))


@app.get("/bullpens", response_class=HTMLResponse)
def bullpens_page() -> HTMLResponse:
    return HTMLResponse(BULLPEN_TEMPLATE_PATH.read_text(encoding="utf-8"))


@app.get("/velocity-comparison", response_class=HTMLResponse)
def velocity_comparison_page() -> HTMLResponse:
    return HTMLResponse(VELOCITY_TEMPLATE_PATH.read_text(encoding="utf-8"))


@app.get("/starting-pitchers", response_class=HTMLResponse)
def starting_pitchers_page() -> HTMLResponse:
    return HTMLResponse(STARTING_PITCHERS_TEMPLATE_PATH.read_text(encoding="utf-8"))


@app.get("/probable-pitchers", response_class=HTMLResponse)
def probable_pitchers_page() -> HTMLResponse:
    return HTMLResponse(PITCHERS_TEMPLATE_PATH.read_text(encoding="utf-8"))


@app.get("/pitchers/{mlb_player_id}", response_class=HTMLResponse)
def pitcher_profile_page(mlb_player_id: int) -> HTMLResponse:
    return HTMLResponse(PITCHER_PROFILE_TEMPLATE_PATH.read_text(encoding="utf-8"))


@app.get("/api/teams", response_class=JSONResponse)
def api_teams() -> JSONResponse:
    return JSONResponse({"teams": TEAM_OPTIONS, "default_team": "braves"})


@app.get("/api/lineups", response_class=JSONResponse)
def api_lineups(team: str = Query("braves", description="Fangraphs team slug")) -> JSONResponse:
    if team not in TEAM_MAP:
        raise HTTPException(status_code=404, detail="Unknown team slug")

    try:
        payload = fetch_lineup_data(team)
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch Fangraphs lineup page: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected parse error: {exc}") from exc

    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/bullpens", response_class=JSONResponse)
def api_bullpens(team: str = Query("braves", description="Fangraphs team slug")) -> JSONResponse:
    if team not in TEAM_MAP:
        raise HTTPException(status_code=404, detail="Unknown team slug")

    try:
        payload = fetch_bullpen_data(team)
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch Fangraphs bullpen page: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected bullpen parse error: {exc}") from exc

    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/velocity-comparison", response_class=JSONResponse)
def api_velocity_comparison() -> JSONResponse:
    try:
        payload = get_or_build_velocity_comparison_cache()
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch Baseball Savant pitch-arsenals data: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected velocity comparison parse error: {exc}") from exc

    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/pitch-stuff-plus", response_class=JSONResponse)
def api_pitch_stuff_plus(
    name: str = Query(..., description="Pitcher name"),
    mlb_id: Optional[int] = Query(None, description="MLBAM pitcher id"),
    fg: Optional[int] = Query(None, description="Fangraphs pitcher id"),
) -> JSONResponse:
    try:
        payload = fetch_pitch_stuff_plus(name, mlb_id=mlb_id, fangraphs_player_id=fg)
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch Fangraphs pitch-level Stuff+ data: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected pitch Stuff+ parse error: {exc}") from exc

    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/probable-pitchers", response_class=JSONResponse)
def api_probable_pitchers(target_date: Optional[str] = Query(None, alias="date", description="Game date in YYYY-MM-DD")) -> JSONResponse:
    try:
        resolved_date = datetime.strptime(target_date, "%Y-%m-%d").date() if target_date else next_mlb_date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="Invalid date. Use YYYY-MM-DD.") from exc

    try:
        payload = fetch_probable_pitchers(resolved_date)
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch MLB scores page: {exc}") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected probable pitchers parse error: {exc}") from exc

    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/starting-pitchers", response_class=JSONResponse)
def api_starting_pitchers() -> JSONResponse:
    try:
        payload = build_starting_pitcher_payload()
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected starting pitcher table error: {exc}") from exc

    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.post("/api/starting-pitchers/override", response_class=JSONResponse)
def api_starting_pitchers_override(
    payload: Dict[str, object] = Body(...),
) -> JSONResponse:
    pitcher_name = str(payload.get("pitcher") or "").strip()
    if not pitcher_name:
        raise HTTPException(status_code=400, detail="Pitcher is required.")

    override_payload = update_starting_pitcher_override(
        pitcher_name=pitcher_name,
        era=payload.get("era"),
        pitches=payload.get("pitches"),
    )
    return JSONResponse(
        {
            "pitcher": pitcher_name,
            "override": override_payload,
        },
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/homepage-scoreboard", response_class=JSONResponse)
def api_homepage_scoreboard(day: Optional[str] = Query(None, description="Optional slate key: today or tomorrow")) -> JSONResponse:
    if day not in (None, "today", "tomorrow"):
        raise HTTPException(status_code=400, detail="Invalid day. Use today or tomorrow.")
    payload = build_homepage_scoreboard_payload(day)
    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/pitchers/{mlb_player_id}", response_class=JSONResponse)
def api_pitcher_profile(mlb_player_id: int, fg: Optional[int] = Query(None, description="Fangraphs player id")) -> JSONResponse:
    try:
        payload = build_pitcher_profile(mlb_player_id, fg)
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch pitcher profile data: {exc}") from exc
    except ValueError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected pitcher profile error: {exc}") from exc

    return JSONResponse(
        payload,
        headers={
            "Cache-Control": "no-store, max-age=0",
            "Pragma": "no-cache",
        },
    )


@app.get("/api/export-lineups.xlsx")
def api_export_lineups() -> StreamingResponse:
    try:
        content = build_excel_workbook()
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch Fangraphs lineup data for export: {exc}") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected export error: {exc}") from exc

    filename = f"mlb_platoon_lineups_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/export-bullpens.xlsx")
def api_export_bullpens() -> StreamingResponse:
    try:
        content = build_bullpen_excel_workbook()
    except (HTTPError, URLError, TimeoutError) as exc:
        raise HTTPException(status_code=502, detail=f"Failed to fetch Fangraphs bullpen data for export: {exc}") from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected bullpen export error: {exc}") from exc

    filename = f"mlb_bullpens_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/health", response_class=JSONResponse)
def health() -> JSONResponse:
    return JSONResponse({"ok": True})


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("platoon_viewer:app", host="127.0.0.1", port=8000, reload=True)
